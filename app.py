import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime, date, time

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="GILDAN — Control de Permisos",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded",
)

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# Hora de fin de turno en la fábrica (para calcular S/R)
HORA_FIN_TURNO = time(15, 0)

PASSWORD = "1234"

PLANTAS  = ["Fábrica San Juan", "Casa Central Bs. As.", "🏢 Total Empresa"]
AÑOS     = list(range(2025, 2036))   # 2025 → 2035

# Tope anual de horas a compensar (año calendario, resetea 1° de enero)
TOPE_HORAS_NORMAL = 8
TOPE_HORAS_LIDER  = 16

# Líderes con tope ampliado — nombres tal como figuran en el padrón (MAYÚSCULA, "APELLIDO, NOMBRE")
# Si cambia el padrón, mantener esta lista sincronizada con el campo es_lider.
LIDERES_SJ = {
    "SANTANA, SANDRA BETTINA",
    "RODRIGUEZ, PATRICIA SOLEDAD",
    "TEJADA, DALINDA MATILDE",
    "FLORES FRIAS, CELIA ROMINA",
    "MURO, LILIANA MABEL",
    "CERDA, CLAUDIA DEL VALLE",
    "OLIVA ZEBALLOS, ANALIA",
}

# ── Motivos canónicos ──────────────────────────────────────────
MOTIVOS_LISTA = [
    "Banco / Cajero",
    "Médico propio",
    "Médico turno mañana",
    "Familiar enfermo",
    "Enfermedad propia / Clínica",
    "Obra social / ANSES",
    "Juzgado / Tribunales",
    "Registro Civil / DNI",
    "Escribanía",
    "Carnet de Conducir",
    "Análisis de sangre",
    "Colegio hijo/a",
    "Escuela hijo/a",
    "Cuidado familiar",
    "Trámite personal",
    "Duelo / Fallecimiento familiar",
    "ART",
    "Otro",
]

# Política San Juan (RR.HH. 020): motivos que PERMITEN compensar
# Para estos el guardia puede elegir SI o NO.
# Para los demás se fuerza NO automáticamente.
MOTIVOS_COMPENSAN_SJ = {
    "Banco / Cajero",
    "Análisis de sangre",
    "Carnet de Conducir",
    "Registro Civil / DNI",
    "Obra social / ANSES",
    "Juzgado / Tribunales",   # política modificada — sí compensa
    "Escribanía",
    "Trámite personal",
    "Colegio hijo/a",
    "Médico turno mañana",    # nuevo: turno mañana sí compensa en SJ
    "Otro",                   # queda a criterio
}

# Política Buenos Aires (RR.HH. 036 — convenio): motivos que PERMITEN compensar
# Contemplados: análisis/estudios médicos (propios e hijos), trámites bancarios,
# trámites en registros/dependencias públicas, reuniones/actos escolares de hijos.
# Máximo 10hs anuales, hasta 4hs por permiso (control operativo, no en la app).
MOTIVOS_COMPENSAN_BSAS = {
    "Banco / Cajero",
    "Análisis de sangre",
    "Médico propio",        # estudios médicos propios
    "Familiar enfermo",     # estudios médicos hijos
    "Registro Civil / DNI", # trámites en dependencias públicas
    "Obra social / ANSES",  # trámites en dependencias públicas
    "Juzgado / Tribunales", # trámites en dependencias públicas
    "Escribanía",           # trámites en dependencias públicas
    "Escuela hijo/a",       # reuniones/actos escolares
    "Colegio hijo/a",       # reuniones/actos escolares
    "Otro",                 # queda a criterio
}

# Mapeo para normalizar motivos históricos escritos a mano
MOTIVO_MAP = {
    "bco":                       "Banco / Cajero",
    "banco":                     "Banco / Cajero",
    "medico":                    "Médico propio",
    "médico":                    "Médico propio",
    "turno médico":              "Médico propio",
    "turno medico":              "Médico propio",
    "junta medica":              "Médico propio",
    "junta médica":              "Médico propio",
    "medico hijos":              "Familiar enfermo",
    "hija enferma":              "Familiar enfermo",
    "retiro hija + emicar":      "Familiar enfermo",
    "enferma":                   "Enfermedad propia",
    "enfermedad":                "Enfermedad propia",
    "presión alta":              "Enfermedad propia",
    "presion alta":              "Enfermedad propia",
    "obra social":               "Obra social / ANSES",
    "anses":                     "Obra social / ANSES",
    "juzgado":                   "Juzgado / Tribunales",
    "tribunales":                "Juzgado / Tribunales",
    "ufi":                       "Juzgado / Tribunales",
    "declarar estafa":           "Juzgado / Tribunales",
    "registro civil":            "Registro Civil / DNI",
    "dni":                       "Registro Civil / DNI",
    "carnet de conducir":        "Carnet de Conducir",
    "escribania":                "Escribanía",
    "escribanía":                "Escribanía",
    "emicar":                    "Carnet de Conducir",
    "sanatorio sj":              "Enfermedad propia / Clínica",
    "clinica":                   "Enfermedad propia / Clínica",
    "clínica":                   "Enfermedad propia / Clínica",
    "analisis":                  "Análisis de sangre",
    "análisis":                  "Análisis de sangre",
    "analisis de sangre":        "Análisis de sangre",
    "ipv":                       "Trámite personal",
    "personal":                  "Trámite personal",
    "1 dia de clases":           "Colegio hijo/a",
    "colegio":                   "Colegio hijo/a",
    "fallecimiento familiar":    "Duelo / Fallecimiento familiar",
}

MESES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def normalizar_motivo(raw: str) -> str:
    """Convierte cualquier motivo crudo al canónico estandarizado."""
    if not raw or pd.isna(raw):
        return "Otro"
    return MOTIVO_MAP.get(str(raw).strip().lower(), str(raw).strip().title())


# ─────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────
def check_login():
    if "autenticado" not in st.session_state:
        st.session_state.autenticado = False
    if not st.session_state.autenticado:
        st.markdown("# 🧵 GILDAN")
        st.title("Control de Permisos — Recursos Humanos")
        st.divider()
        st.write("Ingresá la contraseña para acceder.")
        clave = st.text_input("Contraseña", type="password", placeholder="••••")
        if st.button("Ingresar", use_container_width=True, type="primary"):
            if clave == PASSWORD:
                st.session_state.autenticado = True
                st.rerun()
            else:
                st.error("Contraseña incorrecta.")
        st.stop()

check_login()


# ─────────────────────────────────────────────
# REDONDEO DE HORAS
# ─────────────────────────────────────────────
def redondear_horas(minutos: float) -> float:
    """
    < 30 min  → 0h
    30–89 min → 1h
    90–149    → 2h  (fracción >= :30 sube al entero siguiente)
    """
    if minutos is None or pd.isna(minutos) or minutos < 30:
        return 0.0
    parte = int(minutos // 60)
    fraccion = (minutos % 60) / 60
    return float(parte + 1) if fraccion >= 0.5 else float(parte)


def minutos_entre(t_sal: time, t_ent: time) -> float:
    """Minutos entre dos objetos time (mismo día)."""
    ref = date.today()
    delta = datetime.combine(ref, t_ent) - datetime.combine(ref, t_sal)
    return round(delta.seconds / 60, 1)


def fmt_dur(minutos: float) -> str:
    h = int(minutos // 60)
    m = int(minutos % 60)
    if h == 0:
        return f"{m} min"
    return f"{h}h {m:02d}min" if m else f"{h}h"


def fmt_horas(h: float) -> str:
    """Formatea horas con soporte de medias horas: 0.5 → '0.5h', 1 → '1h', 1.5 → '1.5h'."""
    if h == int(h):
        return f"{int(h)}h"
    return f"{h:.1f}h"


# ─────────────────────────────────────────────
# CONEXIÓN GOOGLE SHEETS
# ─────────────────────────────────────────────
@st.cache_resource(ttl=3600)
def conectar_sheets():
    creds = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"], scopes=SCOPES
    )
    return gspread.authorize(creds)


@st.cache_resource(ttl=3600)
def get_wb(_gc):
    return _gc.open_by_key(st.secrets["SHEET_ID"])


@st.cache_data(ttl=60)
def leer_padron(_gc):
    ws = get_wb(_gc).worksheet("padron")
    df = pd.DataFrame(ws.get_all_records())
    if not df.empty:
        df["legajo"] = df["legajo"].astype(str).str.strip()
        df["nombre"] = df["nombre"].astype(str).str.strip().str.upper()
        df["planta"] = df["planta"].astype(str).str.strip()
        # sector y clasificacion pueden no existir si aún no se corrió el script
        if "sector" not in df.columns:
            df["sector"] = ""
        else:
            df["sector"] = df["sector"].astype(str).str.strip()
        if "clasificacion" not in df.columns:
            df["clasificacion"] = ""
        else:
            df["clasificacion"] = df["clasificacion"].astype(str).str.strip()
        # es_lider: columna nueva en Sheets. Si no existe todavía, se infiere
        # desde LIDERES_SJ como fallback (hasta que se corra el script de migración).
        if "es_lider" not in df.columns:
            df["es_lider"] = df["nombre"].apply(lambda n: "SI" if n in LIDERES_SJ else "NO")
        else:
            df["es_lider"] = df["es_lider"].astype(str).str.strip().str.upper()
            df["es_lider"] = df["es_lider"].where(df["es_lider"].isin(["SI", "NO"]),
                                                    df["nombre"].apply(lambda n: "SI" if n in LIDERES_SJ else "NO"))
    return df


@st.cache_data(ttl=20)
def leer_permisos(_gc):
    ws = get_wb(_gc).worksheet("permisos")
    df = pd.DataFrame(ws.get_all_records())
    if df.empty:
        return df
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    df["minutos_reales"] = pd.to_numeric(df["minutos_reales"], errors="coerce")
    df["horas_redondeadas"] = pd.to_numeric(df["horas_redondeadas"], errors="coerce")
    df["compensa"] = df["compensa"].astype(str).str.upper().str.strip()
    df["legajo"] = df["legajo"].astype(str).str.strip()
    df["planta"] = df["planta"].astype(str).str.strip()
    # Normalizar motivos al leer
    df["motivo"] = df["motivo"].apply(normalizar_motivo)
    return df


@st.cache_data(ttl=20)
def leer_compensaciones(_gc):
    ws = get_wb(_gc).worksheet("compensaciones")
    df = pd.DataFrame(ws.get_all_records())
    if df.empty:
        return df
    df["fecha_compensacion"] = pd.to_datetime(df["fecha_compensacion"], errors="coerce")
    df["horas_compensadas"] = pd.to_numeric(df["horas_compensadas"], errors="coerce")
    df["legajo"] = df["legajo"].astype(str).str.strip()
    df["planta"] = df["planta"].astype(str).str.strip()
    return df


def guardar_permiso(gc, fila: dict):
    get_wb(gc).worksheet("permisos").append_row(list(fila.values()), value_input_option="RAW")
    leer_permisos.clear()


def guardar_compensacion(gc, fila: dict):
    get_wb(gc).worksheet("compensaciones").append_row(list(fila.values()), value_input_option="RAW")
    leer_compensaciones.clear()


def agregar_empleado(gc, legajo: str, nombre: str, sector: str, planta: str,
                      clasificacion: str = "", es_lider: str = "NO"):
    # Schema padron: legajo, nombre, sector, centro_costo, planta, activo, clasificacion, es_lider
    get_wb(gc).worksheet("padron").append_row(
        [legajo.strip(), nombre.upper().strip(), sector.strip(), "", planta, "SI",
         clasificacion.strip(), es_lider],
        value_input_option="RAW"
    )
    leer_padron.clear()


def generar_id(prefijo="P"):
    return f"{prefijo}{datetime.now().strftime('%Y%m%d%H%M%S')}"


# ─────────────────────────────────────────────
# TOPE ANUAL DE COMPENSACIÓN
# ─────────────────────────────────────────────
def obtener_tope(es_lider: str) -> float:
    """Tope anual de horas a compensar según si la persona es líder."""
    return TOPE_HORAS_LIDER if str(es_lider).strip().upper() == "SI" else TOPE_HORAS_NORMAL


def horas_comprometidas_año(permisos_df: pd.DataFrame, legajo: str, año: int) -> float:
    """
    Suma las horas con compensa=SI que una persona ya tiene cargadas
    en el año calendario indicado. Año calendario = resetea 1° de enero.
    """
    if permisos_df.empty:
        return 0.0
    p = permisos_df[
        (permisos_df["legajo"] == legajo) &
        (permisos_df["compensa"] == "SI") &
        (permisos_df["fecha"].dt.year == año)
    ]
    return float(p["horas_redondeadas"].sum()) if not p.empty else 0.0


# ─────────────────────────────────────────────
# CALCULAR SALDOS
# ─────────────────────────────────────────────
def calcular_saldos(
    permisos_df: pd.DataFrame,
    comp_df: pd.DataFrame,
    hasta_fecha=None,
) -> pd.DataFrame:
    """
    Calcula saldo acumulado por persona.
    Si hasta_fecha (date), filtra permisos y compensaciones hasta esa fecha inclusive.
    Retorna solo personas con saldo > 0.
    """
    cols = ["legajo", "nombre", "debe", "compensado", "saldo"]
    if permisos_df.empty:
        return pd.DataFrame(columns=cols)

    p = permisos_df[permisos_df["compensa"] == "SI"].copy()
    c = comp_df.copy() if not comp_df.empty else pd.DataFrame()

    if hasta_fecha is not None:
        p = p[p["fecha"].dt.date <= hasta_fecha]
        if not c.empty and "fecha_compensacion" in c.columns:
            c = c[c["fecha_compensacion"].dt.date <= hasta_fecha]

    if p.empty:
        return pd.DataFrame(columns=cols)

    debe = (
        p.groupby(["legajo", "nombre"])["horas_redondeadas"]
        .sum().reset_index()
        .rename(columns={"horas_redondeadas": "debe"})
    )
    if not c.empty and "horas_compensadas" in c.columns:
        comp = (
            c.groupby("legajo")["horas_compensadas"]
            .sum().reset_index()
            .rename(columns={"horas_compensadas": "compensado"})
        )
        saldo = debe.merge(comp, on="legajo", how="left")
    else:
        saldo = debe.copy()
        saldo["compensado"] = 0.0
    saldo["compensado"] = saldo["compensado"].fillna(0.0)
    saldo["saldo"] = (saldo["debe"] - saldo["compensado"]).round(1)
    return saldo[saldo["saldo"] > 0].sort_values("saldo", ascending=False).reset_index(drop=True)


# ─────────────────────────────────────────────
# CARGAR DATOS
# ─────────────────────────────────────────────
try:
    gc = conectar_sheets()
except Exception as e:
    st.error(f"No se pudo conectar con Google Sheets: {e}")
    st.stop()

try:
    with st.spinner("Cargando datos..."):
        padron         = leer_padron(gc)
        permisos       = leer_permisos(gc)
        compensaciones = leer_compensaciones(gc)
except Exception as e:
    st.error(f"Error al leer datos: {e}")
    st.stop()


# ─────────────────────────────────────────────
# SIDEBAR — selección de planta + navegación
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🧵 GILDAN")
    st.markdown("### Control de Permisos")
    st.caption(f"Hoy: {date.today().strftime('%d/%m/%Y')}")
    st.divider()

    planta_activa = st.selectbox("📍 Planta", PLANTAS)

    st.divider()
    pagina = st.radio(
        "Sección:",
        ["🔵 Panel Guardia", "🟢 Panel RRHH", "📊 Análisis", "📖 Cómo se calcula"],
        label_visibility="collapsed",
    )
    st.divider()
    if st.button("🔒 Cerrar sesión", use_container_width=True):
        st.session_state.autenticado = False
        st.rerun()

# Filtrar datos por planta seleccionada
KEY_PLANTA = "Fábrica" if "San Juan" in planta_activa else ("Casa Central" if "Bs." in planta_activa else "Total")
ES_TOTAL   = KEY_PLANTA == "Total"
ES_SJ      = KEY_PLANTA == "Fábrica"

padron_planta   = padron.copy() if ES_TOTAL else (padron[padron["planta"] == KEY_PLANTA].copy() if not padron.empty else padron)
permisos_planta = permisos.copy() if ES_TOTAL else (permisos[permisos["planta"] == KEY_PLANTA].copy() if not permisos.empty else permisos)
comp_planta     = compensaciones.copy() if ES_TOTAL else (compensaciones[compensaciones["planta"] == KEY_PLANTA].copy() if not compensaciones.empty else compensaciones)

padron_dict     = dict(zip(padron_planta["legajo"], padron_planta["nombre"])) if not padron_planta.empty else {}
nombre_a_legajo = dict(zip(padron_planta["nombre"], padron_planta["legajo"])) if not padron_planta.empty else {}
nombres_lista   = sorted(padron_planta["nombre"].tolist()) if not padron_planta.empty else []

# Solo empleados activos para reportes de saldo y selectboxes del guardia
padron_activos  = padron_planta[padron_planta["activo"].astype(str).str.upper() == "SI"] if not padron_planta.empty else padron_planta
legajos_activos = set(padron_activos["legajo"].tolist())
nombres_lista   = sorted(padron_activos["nombre"].tolist()) if not padron_activos.empty else []
# Para saldos solo contamos permisos/compensaciones de activos
permisos_activos = permisos_planta[permisos_planta["legajo"].isin(legajos_activos)] if not permisos_planta.empty else permisos_planta
comp_activos     = comp_planta[comp_planta["legajo"].isin(legajos_activos)] if not comp_planta.empty else comp_planta
# Dicts para enriquecer reportes con sector y clasificación
sector_dict  = dict(zip(padron_activos["legajo"], padron_activos["sector"])) if not padron_activos.empty else {}
clasif_dict  = dict(zip(padron_activos["legajo"], padron_activos["clasificacion"])) if not padron_activos.empty else {}


# ═══════════════════════════════════════════════════════════════
# PANEL GUARDIA
# ═══════════════════════════════════════════════════════════════
if pagina == "🔵 Panel Guardia":

    st.title("🏭 Hola San Juan" if "San Juan" in planta_activa else "🏭 Hola Bs. As.")
    st.write("Buscá a la persona por nombre y completá los datos del permiso.")
    st.divider()

    # ── Búsqueda por nombre ──
    st.subheader("¿Quién sale?")
    opciones = ["— Seleccioná un nombre —"] + nombres_lista
    nombre_sel = st.selectbox(
        "Nombre completo",
        opciones,
        help="Escribí las primeras letras para filtrar.",
    )

    legajo_resuelto = ""
    nombre_resuelto = ""

    if nombre_sel and nombre_sel != "— Seleccioná un nombre —":
        legajo_resuelto = nombre_a_legajo.get(nombre_sel, "")
        nombre_resuelto = nombre_sel
        ci1, ci2 = st.columns(2)
        ci1.success(f"✅ **{nombre_resuelto}**")
        ci2.info(f"Legajo: **{legajo_resuelto}**" if legajo_resuelto else "Sin legajo")

    with st.expander("🔢 Buscar por legajo (opcional)"):
        leg_manual = st.text_input("Legajo", placeholder="Ej: 2621")
        if leg_manual.strip() in padron_dict:
            nombre_resuelto = padron_dict[leg_manual.strip()]
            legajo_resuelto = leg_manual.strip()
            st.success(f"✅ {nombre_resuelto}")
        elif leg_manual.strip():
            st.warning("Legajo no encontrado.")

    st.divider()
    st.subheader("Datos del permiso")


    # ── Controles FUERA del form (necesitan reactividad inmediata) ──
    # Motivo + campo "Otro" reactivo
    motivo_sel = st.selectbox("📋 Motivo de salida *", MOTIVOS_LISTA, key="motivo_pre")
    motivo_otro = ""
    if motivo_sel == "Otro":
        motivo_otro = st.text_input(
            "✏️ Especificá el motivo",
            placeholder="Ej: Trámite bancario especial, reunión escuela, etc.",
            max_chars=80,
            key="motivo_otro_pre",
        )

    # Hora de salida y Sin retorno — fuera del form para reactividad
    col_sr1, col_sr2 = st.columns([2, 1])
    with col_sr1:
        hora_salida_pre = st.time_input("🚪 Hora de salida *", value=time(8, 0), step=60, key="hora_sal_pre")
    with col_sr2:
        sin_retorno_pre = st.checkbox("🔴 Sin retorno (no volvió)", value=False, key="sr_pre")

    # Política SJ: determinar si el motivo permite compensar
    if ES_SJ:
        puede_compensar = motivo_sel in MOTIVOS_COMPENSAN_SJ
        if not puede_compensar:
            st.info(
                f"ℹ️ El motivo **{motivo_sel}** no está contemplado en la política de compensación "
                "de San Juan — se registrará como **No compensa** automáticamente."
            )
    elif KEY_PLANTA == "Casa Central":
        puede_compensar = motivo_sel in MOTIVOS_COMPENSAN_BSAS
        if not puede_compensar:
            st.info(
                f"ℹ️ El motivo **{motivo_sel}** no está contemplado en la política de "
                "compensación de Bs. As. (RR.HH. 036) — se registrará como **No compensa** automáticamente."
            )
    else:
        # Total Empresa: sin política definida globalmente, queda a criterio
        puede_compensar = True

    # ── TOPE ANUAL — chequeo antes de habilitar el radio SI/NO ──
    # Si la persona ya alcanzó su tope de horas a compensar en el año
    # calendario actual, se fuerza NO sin importar lo que diga la política
    # de motivos. El guardia no puede elegir SI en ese caso.
    tope_alcanzado = False
    if legajo_resuelto and ES_SJ:
        _es_lider_sel = padron_activos[padron_activos["legajo"] == legajo_resuelto]["es_lider"].values
        _es_lider_sel = _es_lider_sel[0] if len(_es_lider_sel) > 0 else "NO"
        _tope_sel = obtener_tope(_es_lider_sel)
        _año_actual = date.today().year
        _comprometidas_sel = horas_comprometidas_año(permisos_activos, legajo_resuelto, _año_actual)

        st.caption(
            f"📊 Tope anual {_año_actual}: **{_comprometidas_sel:.1f}h / {_tope_sel:.0f}h** "
            f"{'(líder)' if _es_lider_sel == 'SI' else ''}"
        )

        if _comprometidas_sel >= _tope_sel:
            tope_alcanzado = True
            puede_compensar = False
            st.warning(
                f"⚠️ **{nombre_resuelto}** ya alcanzó el tope anual de {_tope_sel:.0f}h para compensar "
                f"en {_año_actual} (lleva {_comprometidas_sel:.1f}h). Este permiso se registrará como "
                "**No compensa** automáticamente, sin importar el motivo."
            )

    # Radio compensa — FUERA del form para reactividad con el motivo
    if puede_compensar:
        compensa_pre = st.radio(
            "💰 ¿Va a compensar las horas?",
            ["SI", "NO"],
            horizontal=True,
            key="compensa_pre",
            help="SI = se queda horas extra otro día. NO = no se descuenta.",
        )
    else:
        compensa_pre = "NO"
        if not tope_alcanzado:
            st.write("💰 **No compensa** (automático por política)")

    # Si S/R → hora entrada automática 15:00 y deshabilitada
    valor_entrada = HORA_FIN_TURNO if sin_retorno_pre else time(9, 0)

    with st.form("form_guardia", clear_on_submit=True):

        fecha_permiso = st.date_input("📅 Fecha", value=date.today(), max_value=date.today(), format="DD/MM/YYYY")

        # Leer valores del pre-form
        hora_salida  = hora_salida_pre
        sin_retorno  = sin_retorno_pre
        compensa     = compensa_pre
        hora_entrada = st.time_input(
            "🏁 Hora de entrada" + (" (automático — Sin retorno: 15:00)" if sin_retorno else ""),
            value=valor_entrada,
            step=60,
            disabled=sin_retorno,
        )
        registrado_por = st.text_input("👮 Tu nombre *", placeholder="Ej: García Juan")

        # ── PREVISUALIZACIÓN ──
        # S/R: se calcula contra fin de turno (15:00)
        # Normal: diferencia entre salida y entrada
        if sin_retorno:
            if hora_salida < HORA_FIN_TURNO:
                mins_prev = minutos_entre(hora_salida, HORA_FIN_TURNO)
                hrs_prev  = redondear_horas(mins_prev)
                st.info(
                    f"🔴 Sin retorno — salió a las {hora_salida.strftime('%H:%M')}, "
                    f"fin de turno 15:00 → **{fmt_dur(mins_prev)} real** "
                    f"→ {'**' + str(int(hrs_prev)) + 'h a compensar**' if compensa == 'SI' else 'no compensa'}"
                )
            else:
                st.warning("La hora de salida es posterior al fin de turno (15:00). Verificá.")
        elif hora_entrada > hora_salida:
            mins_prev = minutos_entre(hora_salida, hora_entrada)
            hrs_prev  = redondear_horas(mins_prev)
            st.info(
                f"⏱ Tiempo real fuera: **{fmt_dur(mins_prev)}** "
                f"→ Horas a compensar: **{int(hrs_prev)}h**"
            )

        submitted = st.form_submit_button(
            "💾 GUARDAR REGISTRO", use_container_width=True, type="primary"
        )

        if submitted:
            motivo_final = motivo_otro.strip() if motivo_sel == "Otro" and motivo_otro.strip() else motivo_sel
            errores = []
            if not nombre_resuelto:
                errores.append("Seleccioná o buscá a la persona primero.")
            if not registrado_por.strip():
                errores.append("Falta tu nombre.")
            if sin_retorno and hora_salida >= HORA_FIN_TURNO:
                errores.append("Hora de salida posterior al fin de turno (15:00). Verificá.")
            if not sin_retorno and hora_entrada <= hora_salida:
                errores.append("La hora de entrada debe ser posterior a la salida.")

            if errores:
                for e in errores:
                    st.error(f"❌ {e}")
            else:
                if sin_retorno:
                    # Calcular contra fin de turno
                    mins_r = minutos_entre(hora_salida, HORA_FIN_TURNO)
                    hrs_r  = redondear_horas(mins_r) if compensa == "SI" else 0.0
                    ent_str = "S/R"
                else:
                    mins_r = minutos_entre(hora_salida, hora_entrada)
                    hrs_r  = redondear_horas(mins_r)
                    ent_str = hora_entrada.strftime("%H:%M")

                try:
                    guardar_permiso(gc, {
                        "id":               generar_id("P"),
                        "fecha":            fecha_permiso.strftime("%Y-%m-%d"),
                        "legajo":           legajo_resuelto,
                        "nombre":           nombre_resuelto,
                        "hora_salida":      hora_salida.strftime("%H:%M"),
                        "hora_entrada":     ent_str,
                        "sin_retorno":      "SI" if sin_retorno else "NO",
                        "motivo":           motivo_final,
                        "compensa":         compensa,
                        "minutos_reales":   round(mins_r, 1),
                        "horas_redondeadas": hrs_r,
                        "registrado_por":   registrado_por.strip(),
                        "planta":           KEY_PLANTA,
                        "timestamp":        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    })
                    st.success(f"✅ Registro guardado — {nombre_resuelto}")
                    st.info(
                        f"Tiempo fuera: {fmt_dur(mins_r)} → "
                        f"**{int(hrs_r)}h a compensar**"
                        if compensa == "SI" else
                        f"Tiempo fuera: {fmt_dur(mins_r)} — no compensa"
                    )
                except Exception as e:
                    st.error(f"❌ Error al guardar: {e}")

    # Registros del día
    st.divider()
    st.subheader("Registros de hoy")
    if not permisos_planta.empty:
        hoy = permisos_planta[permisos_planta["fecha"].dt.date == date.today()]
        if hoy.empty:
            st.caption("Aún no hay registros hoy.")
        else:
            hs = hoy[["nombre", "hora_salida", "hora_entrada", "sin_retorno", "motivo", "compensa", "horas_redondeadas"]].copy()
            hs.columns = ["Nombre", "Salida", "Entrada", "S/R", "Motivo", "Compensa", "Hs."]
            st.dataframe(hs, use_container_width=True, hide_index=True)
    else:
        st.caption("No hay registros aún.")

    # Agregar empleado
    st.divider()
    with st.expander("➕ Agregar empleado que no está en la lista"):
        st.caption(
            "Usá esto solo si la persona no aparece en el listado. "
            "El nombre debe ser completo — no puede haber dos personas con el mismo nombre."
        )
        # Listas para los dropdowns — desde el padrón existente
        _sectores_raw = sorted(padron["sector"].dropna().unique().tolist()) if not padron.empty else []
        _sectores     = [s for s in _sectores_raw if s and s not in ("", "nan")]
        _clasifs_raw  = sorted(padron["clasificacion"].dropna().unique().tolist()) if not padron.empty else []
        _clasifs      = [c for c in _clasifs_raw if c and c not in ("", "nan")]
        if not _sectores:
            _sectores = ["COSTURA SAN JUAN", "CALIDAD", "ABASTECIMIENTO", "MANTENIMIENTO",
                         "PLANEAMIENTO", "ADMINISTRACIÓN", "OTRO"]
        if not _clasifs:
            _clasifs = ["HOURLY DIRECT", "HOURLY INDIRECT", "EXEMPT", "NON EXEMPT"]

        with st.form("form_nuevo"):
            cn1, cn2 = st.columns(2)
            with cn1:
                nvo_leg = st.text_input("Legajo *", placeholder="Ej: 3050")
            with cn2:
                nvo_nom = st.text_input("Apellido y Nombre *", placeholder="Ej: GOMEZ, CARLOS ALBERTO")
            na1, na2 = st.columns(2)
            with na1:
                nvo_sec = st.selectbox("Sector *", _sectores)
            with na2:
                nvo_cla = st.selectbox("Clasificación *", _clasifs)
            _planta_opts = ["Fábrica", "Casa Central"]
            _planta_def = 0 if ES_SJ else (1 if not ES_TOTAL else 0)
            nvo_planta = st.selectbox("Planta *", _planta_opts, index=_planta_def)
            nvo_lider = st.radio(
                "¿Es líder? *",
                ["NO", "SI"],
                horizontal=True,
                help=f"Las líderes tienen tope anual de {TOPE_HORAS_LIDER}h para compensar "
                     f"en vez de {TOPE_HORAS_NORMAL}h.",
            )
            if st.form_submit_button("Agregar al padrón", use_container_width=True):
                nom_clean = nvo_nom.strip().upper()
                err = []
                if not nvo_leg.strip():
                    err.append("El legajo es obligatorio.")
                if not nom_clean:
                    err.append("El nombre es obligatorio.")
                if nom_clean in nombre_a_legajo:
                    err.append(f"Ya existe '{nom_clean}'. Agregá segundo nombre o apellido.")
                if err:
                    for e in err:
                        st.error(f"❌ {e}")
                else:
                    try:
                        agregar_empleado(gc, nvo_leg.strip(), nom_clean, nvo_sec,
                                         nvo_planta, nvo_cla, nvo_lider)
                        st.success(f"✅ {nom_clean} agregado. Recargá la página.")
                    except Exception as e:
                        st.error(f"❌ Error: {e}")


# ═══════════════════════════════════════════════════════════════
# PANEL RRHH
# ═══════════════════════════════════════════════════════════════
elif pagina == "🟢 Panel RRHH":

    _rrhh_titulo = ("🟢 RRHH — San Juan" if "San Juan" in planta_activa
                   else ("🟢 RRHH — Bs. As." if "Bs." in planta_activa else "🟢 RRHH — Total Empresa"))
    st.title(_rrhh_titulo)
    st.write("Seguimiento de permisos y compensaciones.")
    st.divider()

    # ── Gestión de empleados (inactivos / baja) ──────────────────
    with st.expander("👤 Gestionar empleados — marcar inactivo o dar de baja"):
        st.caption(
            "Marcá como **Inactivo** a quien se fue de la empresa: sus horas pendientes "
            "desaparecen del reporte pero quedan en el historial. "
            "**Dar de baja** elimina el empleado del padrón definitivamente."
        )
        if not padron_planta.empty:
            emp_opciones = ["— Seleccioná —"] + sorted(padron_planta["nombre"].tolist())
            emp_sel_g = st.selectbox("Empleado/a", emp_opciones, key="emp_gestion")
            if emp_sel_g and emp_sel_g != "— Seleccioná —":
                leg_emp_g = nombre_a_legajo.get(emp_sel_g, "")
                activo_vals = padron_planta[padron_planta["nombre"] == emp_sel_g]["activo"].values
                activo_actual = str(activo_vals[0]).upper() if len(activo_vals) > 0 else "SI"
                st.write(f"Estado actual: **{'✅ Activo' if activo_actual == 'SI' else '⛔ Inactivo'}**")
                col_g1, col_g2 = st.columns(2)
                with col_g1:
                    if activo_actual == "SI":
                        if st.button("⛔ Marcar como Inactivo", use_container_width=True, key="btn_inactivo"):
                            try:
                                ws_p = get_wb(gc).worksheet("padron")
                                for c in ws_p.findall(emp_sel_g):
                                    ws_p.update_cell(c.row, 5, "NO")
                                leer_padron.clear()
                                st.success(f"⛔ {emp_sel_g} marcado como inactivo.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ {e}")
                    else:
                        if st.button("✅ Reactivar", use_container_width=True, key="btn_reactivar"):
                            try:
                                ws_p = get_wb(gc).worksheet("padron")
                                for c in ws_p.findall(emp_sel_g):
                                    ws_p.update_cell(c.row, 5, "SI")
                                leer_padron.clear()
                                st.success(f"✅ {emp_sel_g} reactivado.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ {e}")
                with col_g2:
                    if st.button("🗑️ Dar de baja del padrón", use_container_width=True, key="btn_baja",
                                 type="primary"):
                        try:
                            ws_p = get_wb(gc).worksheet("padron")
                            for c in sorted(ws_p.findall(emp_sel_g), key=lambda x: x.row, reverse=True):
                                ws_p.delete_rows(c.row)
                            leer_padron.clear()
                            st.success(f"🗑️ {emp_sel_g} eliminado del padrón.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ {e}")
        else:
            st.info("No hay empleados para esta planta.")

    st.divider()
    col_f1, col_f2, col_f3 = st.columns([1, 1, 2])
    with col_f1:
        año_sel = st.selectbox("Año", AÑOS, index=AÑOS.index(min(date.today().year, max(AÑOS))))
    with col_f2:
        mes_sel = st.selectbox(
            "Mes", list(MESES.keys()),
            index=datetime.now().month - 1,
            format_func=lambda x: MESES[x],
        )
    with col_f3:
        modo = st.radio("Vista:", ["Solo este mes", "Saldo acumulado total"], horizontal=True)

    if permisos_planta.empty:
        st.warning("No hay permisos cargados para esta planta.")
        st.stop()

    if modo == "Solo este mes":
        p_f = permisos_activos[
            (permisos_activos["fecha"].dt.year == año_sel) &
            (permisos_activos["fecha"].dt.month == mes_sel)
        ].copy()
        c_f = comp_activos[
            (comp_activos["fecha_compensacion"].dt.year == año_sel) &
            (comp_activos["fecha_compensacion"].dt.month == mes_sel)
        ].copy() if not comp_activos.empty else comp_activos.copy()
    else:
        p_f = permisos_activos.copy()
        c_f = comp_activos.copy()

    import calendar as _cal
    _ultimo_dia = date(año_sel, mes_sel, _cal.monthrange(año_sel, mes_sel)[1])

    if modo == "Solo este mes":
        # Reporte: horas generadas en este mes → al cierre de ese mes
        saldos_al_cierre = calcular_saldos(
            permisos_activos[
                (permisos_activos["fecha"].dt.year == año_sel) &
                (permisos_activos["fecha"].dt.month == mes_sel)
            ],
            comp_activos,   # compensaciones de cualquier fecha que apliquen
            hasta_fecha=_ultimo_dia,
        )
    else:
        # Reporte: saldo acumulado total hasta hoy
        saldos_al_cierre = calcular_saldos(permisos_activos, comp_activos)

    saldos_actuales  = calcular_saldos(permisos_activos, comp_activos)
    saldos           = saldos_actuales  # alias para métricas
    comp_total = c_f["horas_compensadas"].sum() if not c_f.empty and "horas_compensadas" in c_f.columns else 0

    # Métricas — p_f y c_f ya están filtrados por modo (mes o total)
    _label_periodo = f"{MESES[mes_sel]} {año_sel}" if modo == "Solo este mes" else "Total acumulado"
    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
    col_m1.metric(f"📋 Permisos ({_label_periodo})", len(p_f))
    col_m2.metric("👥 Personas con saldo > 0", len(saldos_al_cierre))
    col_m3.metric("⏳ Horas pendientes (hoy)", f"{saldos_actuales['saldo'].sum():.0f}h" if not saldos_actuales.empty else "0h")
    col_m4.metric(f"✅ Hs. compensadas ({_label_periodo})", f"{comp_total:.0f}h")

    # ── Reporte para gerencia ──────────────────────────────────
    # Incluye Base anual (8h normal / 16h líder), horas consumidas en el
    # año calendario, disponible y semáforo de estado. Reemplaza el
    # proceso manual en Excel que se hacía antes de esta función.
    st.divider()
    st.subheader(f"📄 Reporte para Gerencia — {MESES[mes_sel]} {año_sel}")
    st.caption(
        "Quiénes deben compensar horas, base anual, consumido y disponible. "
        "Esta es la ÚNICA fuente del reporte — descargalo en PNG, no lo copies a mano."
    )
    if saldos_al_cierre.empty:
        st.success(f"✅ Ningún empleado tiene horas pendientes al cierre de {MESES[mes_sel]} {año_sel}.")
    else:
        _año_rep = año_sel  # año calendario sobre el que se evalúa el tope

        # Enriquecer saldos con sector, clasificación, líder, tope y consumido del año
        rep = saldos_al_cierre.copy()
        rep["sector"]        = rep["legajo"].map(sector_dict).fillna("Sin sector")
        rep["clasificacion"] = rep["legajo"].map(clasif_dict).fillna("Sin clasificar")
        rep["es_lider"]      = rep["legajo"].map(
            dict(zip(padron_activos["legajo"], padron_activos["es_lider"]))
        ).fillna("NO")
        rep["base_anual"]    = rep["es_lider"].apply(obtener_tope)
        rep["consumido_año"] = rep["legajo"].apply(
            lambda leg: horas_comprometidas_año(permisos_activos, leg, _año_rep)
        )
        rep["disponible"]    = (rep["base_anual"] - rep["consumido_año"]).clip(lower=0)
        rep["excedente"]     = (rep["consumido_año"] - rep["base_anual"]).clip(lower=0)

        def _estado(row):
            if row["excedente"] > 0:
                return "🔴 EXCEDE TOPE"
            if row["consumido_año"] >= row["base_anual"]:
                return "🔴 LÍMITE"
            if row["disponible"] <= 2:
                return "🟡 ATENCIÓN"
            return "🟢 OK"

        rep["estado"] = rep.apply(_estado, axis=1)
        rep = rep.sort_values(["sector", "clasificacion", "saldo"], ascending=[True, True, False])

        # Colores por clasificación
        CLASIF_COLOR = {
            "HOURLY DIRECT":   "#EBF5FB",
            "HOURLY INDIRECT": "#EAF4F4",
            "EXEMPT":          "#FEF9E7",
            "NON EXEMPT":      "#FDEDEC",
        }

        # Tabla HTML con agrupación por sector
        sector_actual = None
        html_rows = []
        for _, row in rep.iterrows():
            if row["sector"] != sector_actual:
                sector_actual = row["sector"]
                html_rows.append(
                    f'<tr style="background:#1B4F9B;color:white;font-weight:700;">'
                    f'<td colspan="7" style="padding:6px 10px;">📁 {sector_actual}</td></tr>'
                )
            bg = CLASIF_COLOR.get(row["clasificacion"], "#F8F9FA")
            _nombre_disp = row["nombre"] + (" 👑" if row["es_lider"] == "SI" else "")
            html_rows.append(
                f'<tr style="background:{bg};">'
                f'<td style="padding:5px 10px;">{_nombre_disp}</td>'
                f'<td style="padding:5px 10px;color:#555;font-size:0.85rem;">{row["clasificacion"]}</td>'
                f'<td style="padding:5px 10px;text-align:center;">{fmt_horas(row["base_anual"])}</td>'
                f'<td style="padding:5px 10px;text-align:center;">{fmt_horas(row["saldo"])}</td>'
                f'<td style="padding:5px 10px;text-align:center;">{fmt_horas(row["consumido_año"])}</td>'
                f'<td style="padding:5px 10px;font-weight:700;color:#1A7A4A;text-align:center;">'
                f'{fmt_horas(row["disponible"])}</td>'
                f'<td style="padding:5px 10px;text-align:center;">{row["estado"]}</td>'
                f'</tr>'
            )

        tabla_html = f"""
        <style>
            .rep-table {{width:100%;border-collapse:collapse;font-size:0.85rem;font-family:sans-serif;}}
            .rep-table th {{background:#1B4F9B;color:white;padding:8px 8px;text-align:left;font-size:0.78rem;}}
            .rep-table tr:hover td {{filter:brightness(0.96);}}
        </style>
        <table class="rep-table">
          <thead><tr>
            <th>Apellido y Nombre</th>
            <th>Clasificación</th>
            <th style="text-align:center;">Base anual</th>
            <th style="text-align:center;">Pendiente</th>
            <th style="text-align:center;">Consumido {_año_rep}</th>
            <th style="text-align:center;">Disponible</th>
            <th style="text-align:center;">Estado</th>
          </tr></thead>
          <tbody>{"".join(html_rows)}</tbody>
        </table>
        """
        st.markdown(tabla_html, unsafe_allow_html=True)
        st.caption("👑 = líder (tope 16h/año)")
        st.markdown("<br>", unsafe_allow_html=True)

        # Leyenda de colores por clasificación
        col_l1, col_l2, col_l3, col_l4 = st.columns(4)
        for col, (clasif, color) in zip([col_l1, col_l2, col_l3, col_l4], CLASIF_COLOR.items()):
            col.markdown(
                f'<div style="background:{color};border-radius:4px;padding:4px 8px;'
                f'font-size:0.78rem;text-align:center;border:1px solid #ddd;">{clasif}</div>',
                unsafe_allow_html=True
            )

        st.markdown("<br>", unsafe_allow_html=True)
        total_hs = rep["saldo"].sum()
        n_excedentes = (rep["excedente"] > 0).sum()
        st.caption(
            f"**{len(rep)} personas** — **{total_hs:.0f}h** pendientes al cierre de {MESES[mes_sel]}. "
        )
        if n_excedentes > 0:
            st.error(
                f"🔴 **{n_excedentes} persona/s** superaron el tope anual de compensación. "
                "Las horas excedentes no pueden registrarse como compensa=SI a partir de ahora "
                "— el sistema las fuerza a No compensa automáticamente."
            )

        # Botón de descarga PNG — renderiza la tabla completa como imagen
        _png_rep = rep[["nombre","sector","clasificacion","base_anual","saldo",
                        "consumido_año","disponible","estado"]].copy()
        for _col_h in ["base_anual", "saldo", "consumido_año", "disponible"]:
            _png_rep[_col_h] = _png_rep[_col_h].apply(fmt_horas)
        _png_rep.columns = ["Apellido y Nombre","Sector","Clasificación","Base anual",
                            "Pendiente",f"Consumido {_año_rep}","Disponible","Estado"]

        _CLASIF_TEXT = {
            "HOURLY DIRECT":   "#1B4F9B",
            "HOURLY INDIRECT": "#1A7A4A",
            "EXEMPT":          "#7D6608",
            "NON EXEMPT":      "#922B21",
        }
        _fill_colors = []
        _font_colors = []
        for col_name in _png_rep.columns:
            if col_name == "Clasificación":
                _fill_colors.append([CLASIF_COLOR.get(v, "#F8F9FA") for v in _png_rep["Clasificación"]])
                _font_colors.append([_CLASIF_TEXT.get(v, "#333") for v in _png_rep["Clasificación"]])
            elif col_name == "Disponible":
                _fill_colors.append(["#E8F5EE"] * len(_png_rep))
                _font_colors.append(["#1A7A4A"] * len(_png_rep))
            elif col_name == "Pendiente":
                _fill_colors.append(["#FDEDEC"] * len(_png_rep))
                _font_colors.append(["#C0392B"] * len(_png_rep))
            else:
                _fill_colors.append(["white"] * len(_png_rep))
                _font_colors.append(["#222"] * len(_png_rep))

        # PNG — alto calculado correctamente para que no se corte
        _n_filas_png = len(_png_rep)
        _row_h_px    = 32   # px por fila (generoso para que entre texto)
        _header_h    = 40
        _title_h     = 70
        _png_height  = _title_h + _header_h + (_n_filas_png * _row_h_px) + 40

        _fig_png = go.Figure(go.Table(
            header=dict(
                values=list(_png_rep.columns),
                fill_color="#1B4F9B",
                font=dict(color="white", size=11, family="Arial"),
                align="left",
                height=_header_h,
            ),
            cells=dict(
                values=[_png_rep[c] for c in _png_rep.columns],
                fill_color=_fill_colors,
                font=dict(color=_font_colors, size=10, family="Arial"),
                align="left",
                height=_row_h_px,
            ),
        ))
        _fig_png.update_layout(
            title=dict(
                text=f"GILDAN — Reporte de Compensación Horaria<br>"
                     f"<sup>{MESES[mes_sel]} {año_sel} | {planta_activa} | "
                     f"Actualizado al {date.today().strftime('%d/%m/%Y')}</sup>",
                font=dict(size=13, color="#1B4F9B"),
                x=0,
            ),
            margin=dict(t=_title_h, b=30, l=10, r=10),
            height=_png_height,
            width=1100,
        )

        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            try:
                _img_bytes = _fig_png.to_image(format="png", scale=2)
                st.download_button(
                    "⬇️ Descargar PNG",
                    data=_img_bytes,
                    file_name=f"reporte_gerencia_{MESES[mes_sel]}_{año_sel}.png",
                    mime="image/png",
                    use_container_width=True,
                )
            except Exception:
                st.caption("PNG no disponible (kaleido no instalado)")

        with col_dl2:
            # Excel con formato — openpyxl
            try:
                import io as _io
                import openpyxl
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                _wb_xl = openpyxl.Workbook()
                _ws_xl = _wb_xl.active
                _ws_xl.title = f"Reporte {MESES[mes_sel]} {año_sel}"

                # Título
                _ws_xl.merge_cells("A1:H1")
                _ws_xl["A1"] = f"GILDAN — Reporte de Compensación Horaria | {MESES[mes_sel]} {año_sel} | {planta_activa}"
                _ws_xl["A1"].font = Font(bold=True, size=13, color="FFFFFF")
                _ws_xl["A1"].fill = PatternFill("solid", fgColor="1B4F9B")
                _ws_xl["A1"].alignment = Alignment(horizontal="left", vertical="center")
                _ws_xl.row_dimensions[1].height = 28

                # Headers
                _xl_cols = list(_png_rep.columns)
                for ci, col_name in enumerate(_xl_cols, 1):
                    cell = _ws_xl.cell(row=2, column=ci, value=col_name)
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                    cell.fill = PatternFill("solid", fgColor="2471D5")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                _ws_xl.row_dimensions[2].height = 22

                # Colores de fill por clasificación (openpyxl no acepta #)
                _xl_clasif_fill = {
                    "HOURLY DIRECT":   "EBF5FB",
                    "HOURLY INDIRECT": "EAF4F4",
                    "EXEMPT":          "FEF9E7",
                    "NON EXEMPT":      "FDEDEC",
                }
                _xl_estado_fill = {
                    "🟢 OK":         "D5EFE3",
                    "🟡 ATENCIÓN":   "FEF0E0",
                    "🔴 LÍMITE":     "FDEDEC",
                    "🔴 EXCEDE TOPE":"FDEDEC",
                }

                _thin = Side(style="thin", color="CCCCCC")
                _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

                sector_xl = None
                xl_row = 3
                for _, xl_r in _png_rep.iterrows():
                    # Fila separadora de sector
                    if xl_r["Sector"] != sector_xl:
                        sector_xl = xl_r["Sector"]
                        _ws_xl.merge_cells(
                            start_row=xl_row, start_column=1,
                            end_row=xl_row, end_column=len(_xl_cols)
                        )
                        _sec_cell = _ws_xl.cell(row=xl_row, column=1, value=f"  {sector_xl}")
                        _sec_cell.font = Font(bold=True, color="FFFFFF", size=10)
                        _sec_cell.fill = PatternFill("solid", fgColor="1B4F9B")
                        _sec_cell.alignment = Alignment(vertical="center")
                        _ws_xl.row_dimensions[xl_row].height = 18
                        xl_row += 1

                    _clasif_fill = _xl_clasif_fill.get(xl_r.get("Clasificación", ""), "FFFFFF")
                    for ci, col_name in enumerate(_xl_cols, 1):
                        val = xl_r[col_name]
                        cell = _ws_xl.cell(row=xl_row, column=ci, value=val)
                        cell.border = _border
                        cell.alignment = Alignment(horizontal="center" if ci > 2 else "left",
                                                   vertical="center")
                        cell.font = Font(size=10)
                        # Color por columna
                        if col_name == "Clasificación":
                            cell.fill = PatternFill("solid", fgColor=_clasif_fill)
                        elif col_name == "Estado":
                            _ef = _xl_estado_fill.get(str(val), "FFFFFF")
                            cell.fill = PatternFill("solid", fgColor=_ef)
                            cell.font = Font(bold=True, size=10)
                        elif col_name == "Pendiente":
                            cell.fill = PatternFill("solid", fgColor="FDEDEC")
                            cell.font = Font(color="C0392B", bold=True, size=10)
                        elif col_name == "Disponible":
                            cell.fill = PatternFill("solid", fgColor="E8F5EE")
                            cell.font = Font(color="1A7A4A", bold=True, size=10)
                        else:
                            cell.fill = PatternFill("solid", fgColor="FFFFFF")
                    _ws_xl.row_dimensions[xl_row].height = 17
                    xl_row += 1

                # Ancho de columnas
                _col_widths = [32, 22, 18, 11, 11, 14, 11, 16]
                for ci, w in enumerate(_col_widths, 1):
                    _ws_xl.column_dimensions[get_column_letter(ci)].width = w

                _xl_bytes = _io.BytesIO()
                _wb_xl.save(_xl_bytes)
                _xl_bytes.seek(0)
                st.download_button(
                    "⬇️ Descargar Excel",
                    data=_xl_bytes.getvalue(),
                    file_name=f"reporte_gerencia_{MESES[mes_sel]}_{año_sel}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as _xl_err:
                # Fallback CSV
                _csv_fb = _png_rep.copy()
                st.download_button(
                    "⬇️ Descargar CSV",
                    _csv_fb.to_csv(index=False, encoding="utf-8-sig"),
                    file_name=f"reporte_gerencia_{MESES[mes_sel]}_{año_sel}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

    # ── Reporte: personas que NO compensan (para prima de producción) ──
    st.divider()
    st.subheader(f"❌ Reporte: permisos SIN compensación — {MESES[mes_sel]} {año_sel}")
    st.caption(
        "Personas que eligieron **NO compensar** sus horas en el período. "
        "Dato necesario para evaluar prima de producción."
    )

    _p_no_comp = p_f[p_f["compensa"] == "NO"].copy() if not p_f.empty else pd.DataFrame()

    if _p_no_comp.empty:
        st.success("✅ No hay permisos sin compensación en este período.")
    else:
        _p_no_comp["sector_nc"] = _p_no_comp["legajo"].map(sector_dict).fillna("Sin sector")
        _p_no_comp["clasif_nc"] = _p_no_comp["legajo"].map(clasif_dict).fillna("Sin clasificar")
        _p_no_comp["fecha_str"] = _p_no_comp["fecha"].dt.strftime("%d/%m/%Y")

        # Calcular horas reales desde hora_salida/hora_entrada cuando
        # horas_redondeadas = 0 (registros históricos sin minutos_reales guardados)
        def _calc_hs_no_comp(row):
            # Si ya tiene horas_redondeadas válidas, usarlas
            if pd.notna(row["horas_redondeadas"]) and row["horas_redondeadas"] > 0:
                return float(row["horas_redondeadas"])
            # Si tiene minutos_reales, calcular desde ahí
            if pd.notna(row["minutos_reales"]) and row["minutos_reales"] > 0:
                return redondear_horas(row["minutos_reales"])
            # Último recurso: calcular desde hora_salida y hora_entrada (strings "HH:MM")
            try:
                sal = row["hora_salida"]
                ent = row["hora_entrada"]
                if (isinstance(sal, str) and isinstance(ent, str)
                        and len(sal) == 5 and len(ent) == 5 and ent != "S/R"):
                    h_sal, m_sal = map(int, sal.split(":"))
                    h_ent, m_ent = map(int, ent.split(":"))
                    mins = (h_ent * 60 + m_ent) - (h_sal * 60 + m_sal)
                    return redondear_horas(mins) if mins > 0 else 0.0
            except Exception:
                pass
            return 0.0

        _p_no_comp["hs_real"] = _p_no_comp.apply(_calc_hs_no_comp, axis=1)

        _rep_nc = (
            _p_no_comp.groupby(["nombre", "sector_nc", "clasif_nc"])
            .agg(
                permisos=("fecha_str", "count"),
                horas_no_comp=("hs_real", "sum"),
            )
            .reset_index()
            .sort_values(["sector_nc", "horas_no_comp"], ascending=[True, False])
        )
        # Excluir personas con 0 horas calculadas (sin datos de horario)
        _rep_nc = _rep_nc[_rep_nc["horas_no_comp"] > 0].copy()
        _rep_nc["horas_no_comp"] = _rep_nc["horas_no_comp"].apply(fmt_horas)

        # Tabla HTML mismo formato que reporte gerencia
        _CLASIF_COLOR_NC = {
            "HOURLY DIRECT":   "#EBF5FB",
            "HOURLY INDIRECT": "#EAF4F4",
            "EXEMPT":          "#FEF9E7",
            "NON EXEMPT":      "#FDEDEC",
        }
        _sector_act_nc = None
        _html_nc = []
        for _, row in _rep_nc.iterrows():
            if row["sector_nc"] != _sector_act_nc:
                _sector_act_nc = row["sector_nc"]
                _html_nc.append(
                    f'<tr style="background:#922B21;color:white;font-weight:700;">'
                    f'<td colspan="4" style="padding:6px 10px;">📁 {_sector_act_nc}</td></tr>'
                )
            _bg_nc = _CLASIF_COLOR_NC.get(row["clasif_nc"], "#F8F9FA")
            _html_nc.append(
                f'<tr style="background:{_bg_nc};">'
                f'<td style="padding:5px 10px;">{row["nombre"]}</td>'
                f'<td style="padding:5px 10px;color:#555;font-size:0.85rem;">{row["clasif_nc"]}</td>'
                f'<td style="padding:5px 10px;text-align:center;">{row["permisos"]}</td>'
                f'<td style="padding:5px 10px;font-weight:700;color:#922B21;text-align:center;">'
                f'{row["horas_no_comp"]}</td>'
                f'</tr>'
            )
        _tabla_nc_html = f"""
        <style>.nc-table{{width:100%;border-collapse:collapse;font-size:0.9rem;font-family:sans-serif;}}
        .nc-table th{{background:#922B21;color:white;padding:8px 10px;text-align:left;}}</style>
        <table class="nc-table">
          <thead><tr>
            <th>Apellido y Nombre</th><th>Clasificación</th>
            <th style="text-align:center;">Permisos</th>
            <th style="text-align:center;">Hs. no compensadas</th>
          </tr></thead>
          <tbody>{"".join(_html_nc)}</tbody>
        </table>"""
        st.markdown(_tabla_nc_html, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.caption(
            f"**{len(_rep_nc)} personas** eligieron no compensar en {MESES[mes_sel]} {año_sel}."
        )

        # PNG descarga reporte no compensan
        try:
            _fig_nc_png = go.Figure(go.Table(
                header=dict(
                    values=["Apellido y Nombre","Sector","Clasificación","Permisos","Hs. no compensadas"],
                    fill_color="#922B21",
                    font=dict(color="white", size=12, family="Arial"),
                    align="left", height=32,
                ),
                cells=dict(
                    values=[
                        _rep_nc["nombre"], _rep_nc["sector_nc"],
                        _rep_nc["clasif_nc"], _rep_nc["permisos"],
                        _rep_nc["horas_no_comp"],
                    ],
                    fill_color=[
                        ["white"]*len(_rep_nc), ["white"]*len(_rep_nc),
                        [_CLASIF_COLOR_NC.get(v,"#F8F9FA") for v in _rep_nc["clasif_nc"]],
                        ["white"]*len(_rep_nc),
                        ["#FDEDEC"]*len(_rep_nc),
                    ],
                    font=dict(color="#222", size=11, family="Arial"),
                    align="left", height=28,
                ),
            ))
            _fig_nc_png.update_layout(
                title=dict(
                    text=f"GILDAN — Permisos Sin Compensación<br>"
                         f"<sup>{MESES[mes_sel]} {año_sel} | {planta_activa}</sup>",
                    font=dict(size=14, color="#922B21"), x=0,
                ),
                margin=dict(t=70, b=20, l=10, r=10),
                height=max(200, 60 + len(_rep_nc) * 28),
                width=800,
            )
            _nc_bytes = _fig_nc_png.to_image(format="png", scale=2)
            st.download_button(
                "⬇️ Descargar reporte sin compensación (PNG)",
                data=_nc_bytes,
                file_name=f"no_compensan_{MESES[mes_sel]}_{año_sel}.png",
                mime="image/png",
                key="dl_nc_png",
            )
        except Exception:
            pass

    # ── Saldo actual (a hoy) ────────────────────────────────────
    st.divider()
    st.subheader("📊 Saldo acumulado actual (a hoy)")
    st.caption("Incluye todas las compensaciones registradas, incluso de meses futuros.")
    if saldos_actuales.empty:
        st.success("✅ Nadie tiene horas pendientes a la fecha.")
    else:
        sa = saldos_actuales[["nombre","debe","compensado","saldo"]].copy()
        sa["debe"] = sa["debe"].apply(fmt_horas)
        sa["compensado"] = sa["compensado"].apply(fmt_horas)
        sa["saldo"] = saldos_actuales["saldo"].apply(fmt_horas)
        sa.columns = ["Nombre","Debe total","Ya compensó","Saldo hoy"]
        sa.index = range(1, len(sa)+1)
        st.dataframe(sa, use_container_width=True, height=min(380, 45+len(sa)*35))

    # ── Histórico de compensaciones por persona ──────────────────
    st.divider()
    st.subheader("📜 Histórico de compensaciones")
    st.caption(
        "Consultá todas las compensaciones registradas para una persona específica, "
        "incluido el detalle de fecha y observación de cada una."
    )
    _hist_sel = st.selectbox(
        "Empleado/a",
        ["— Seleccioná un nombre —"] + nombres_lista,
        key="hist_comp_sel",
    )
    if _hist_sel and _hist_sel != "— Seleccioná un nombre —":
        _leg_hist = nombre_a_legajo.get(_hist_sel, "")
        _hist_comp = comp_activos[comp_activos["legajo"] == _leg_hist].copy() if not comp_activos.empty else pd.DataFrame()
        _hist_perm = permisos_activos[
            (permisos_activos["legajo"] == _leg_hist) & (permisos_activos["compensa"] == "SI")
        ].copy() if not permisos_activos.empty else pd.DataFrame()

        _es_lider_hist = padron_activos[padron_activos["legajo"] == _leg_hist]["es_lider"].values
        _es_lider_hist = _es_lider_hist[0] if len(_es_lider_hist) > 0 else "NO"
        _tope_hist = obtener_tope(_es_lider_hist)

        col_h1, col_h2, col_h3 = st.columns(3)
        col_h1.metric("Tipo", "👑 Líder" if _es_lider_hist == "SI" else "Empleada")
        col_h1.caption(f"Tope anual: {_tope_hist:.0f}h")
        _consumido_actual_año = horas_comprometidas_año(permisos_activos, _leg_hist, date.today().year)
        col_h2.metric(f"Consumido {date.today().year}", fmt_horas(_consumido_actual_año))
        col_h3.metric("Compensado histórico (todas las fechas)",
                      fmt_horas(_hist_comp["horas_compensadas"].sum()) if not _hist_comp.empty else "0h")

        st.markdown("**Permisos que generaron deuda (compensa = SI)**")
        if _hist_perm.empty:
            st.caption("Sin permisos con compensa=SI registrados.")
        else:
            _hp = _hist_perm[["fecha","motivo","horas_redondeadas"]].copy()
            _hp["fecha"] = _hp["fecha"].dt.strftime("%d/%m/%Y")
            _hp["horas_redondeadas"] = _hp["horas_redondeadas"].apply(fmt_horas)
            _hp.columns = ["Fecha","Motivo","Horas generadas"]
            st.dataframe(_hp.sort_values("Fecha"), use_container_width=True, hide_index=True)

        st.markdown("**Compensaciones registradas**")
        if _hist_comp.empty:
            st.caption("Sin compensaciones registradas todavía.")
        else:
            _hc = _hist_comp[["fecha_compensacion","horas_compensadas","observacion","registrado_por"]].copy()
            _hc["fecha_compensacion"] = _hc["fecha_compensacion"].dt.strftime("%d/%m/%Y")
            _hc["horas_compensadas"] = _hc["horas_compensadas"].apply(fmt_horas)
            _hc.columns = ["Fecha compensó","Horas","Observación","Registrado por"]
            st.dataframe(_hc.sort_values("Fecha compensó"), use_container_width=True, hide_index=True)

    # ── Registrar compensación ──
    st.divider()
    st.subheader("✏️ Registrar compensación")
    st.caption("Cuando alguien se queda horas extra para compensar.")

    with st.form("form_comp"):
        nombre_comp_sel = st.selectbox(
            "Empleado/a",
            ["— Seleccioná un nombre —"] + nombres_lista,
        )
        nom_c = nombre_comp_sel if nombre_comp_sel != "— Seleccioná un nombre —" else ""
        leg_c = nombre_a_legajo.get(nom_c, "") if nom_c else ""

        if nom_c and not saldos.empty:
            saldo_actual = saldos_actuales[saldos_actuales["nombre"] == nom_c]["saldo"].sum() if not saldos_actuales.empty else 0
            if saldo_actual > 0:
                st.info(f"Saldo pendiente de **{nom_c}**: **{saldo_actual:.0f}h**")
            else:
                st.success(f"**{nom_c}** no tiene horas pendientes en el período actual.")

        cc3, cc4 = st.columns(2)
        with cc3:
            fecha_comp = st.date_input("Fecha en que compensó", value=date.today(), format="DD/MM/YYYY")
        with cc4:
            hs_comp = st.number_input("Horas compensadas", min_value=0.5, max_value=8.0, value=1.0, step=0.5)

        obs     = st.text_input("Observación (opcional)", placeholder="Ej: se quedó al final del turno")
        registra = st.text_input("Tu nombre *")

        if st.form_submit_button("✅ REGISTRAR COMPENSACIÓN", use_container_width=True, type="primary"):
            if not nom_c:
                st.error("❌ Seleccioná a la persona.")
            elif not registra.strip():
                st.error("❌ Falta tu nombre.")
            else:
                # Validación: bloquear compensación duplicada (misma persona, mismo día)
                _dup = False
                if not comp_activos.empty and leg_c:
                    _dup_mask = (
                        (comp_activos["legajo"] == leg_c) &
                        (comp_activos["fecha_compensacion"].dt.date == fecha_comp)
                    )
                    if _dup_mask.any():
                        st.error(
                            f"❌ **{nom_c}** ya tiene una compensación registrada "
                            f"el {fecha_comp.strftime('%d/%m/%Y')}. "
                            "No se puede compensar dos veces el mismo día."
                        )
                        _dup = True
                if not _dup:
                    try:
                        guardar_compensacion(gc, {
                            "id":                generar_id("C"),
                            "fecha_compensacion": fecha_comp.strftime("%Y-%m-%d"),
                            "legajo":            leg_c,
                            "nombre":            nom_c,
                            "horas_compensadas": hs_comp,
                            "observacion":       obs,
                            "registrado_por":    registra.strip(),
                            "planta":            KEY_PLANTA,
                            "timestamp":         datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        })
                        st.success(f"✅ {nom_c} — {hs_comp}h el {fecha_comp.strftime('%d/%m/%Y')}")
                    except Exception as e:
                        st.error(f"❌ Error: {e}")

    # Detalle del período
    st.divider()
    st.subheader("🔍 Detalle de permisos del período")
    if not p_f.empty:
        det = p_f[[
            "fecha", "legajo", "nombre", "hora_salida", "hora_entrada",
            "sin_retorno", "motivo", "compensa", "horas_redondeadas"
        ]].copy()
        det["fecha"] = det["fecha"].dt.strftime("%d/%m/%Y")
        det.columns = ["Fecha", "Legajo", "Nombre", "Salida", "Entrada", "S/R", "Motivo", "Compensa", "Hs."]
        st.dataframe(det, use_container_width=True, hide_index=True)
    else:
        st.info("No hay permisos en este período.")


# ═══════════════════════════════════════════════════════════════
# ANÁLISIS
# ═══════════════════════════════════════════════════════════════
elif pagina == "📊 Análisis":

    st.title(f"📊 Análisis de Permisos — {planta_activa}")
    st.caption("Datos históricos completos.")

    if permisos_planta.empty:
        st.warning("No hay datos para analizar en esta planta.")
        st.stop()

    df = permisos_planta.copy()
    df["año"]       = df["fecha"].dt.year
    df["mes"]       = df["fecha"].dt.month
    df["año_semana"] = df["fecha"].dt.strftime("%Y-S%V")

    años_disp = sorted(df["año"].dropna().unique().tolist())
    año_a = st.multiselect("Filtrar por año", años_disp, default=años_disp)
    df = df[df["año"].isin(año_a)]

    if df.empty:
        st.info("No hay datos para ese período.")
        st.stop()

    st.divider()

    # ── 1. Permisos por semana ──
    st.subheader("Permisos por semana")
    sem = df.groupby("año_semana").size().reset_index(name="cantidad")
    prom = sem["cantidad"].mean()
    max_sem = sem["cantidad"].max()

    # Color especial para la semana con más permisos
    colores_sem = ["#C0392B" if v == max_sem else "#1B4F9B" for v in sem["cantidad"]]

    fig1 = go.Figure(go.Bar(
        x=sem["año_semana"],
        y=sem["cantidad"],
        text=sem["cantidad"],
        textposition="outside",
        marker_color=colores_sem,
        hovertemplate="%{x}: %{y} permisos<extra></extra>",
    ))
    fig1.add_hline(
        y=prom, line_dash="dash", line_color="#7F8C8D",
        annotation_text=f"Promedio: {prom:.1f}",
        annotation_position="top left",
    )
    fig1.update_layout(
        plot_bgcolor="white", height=280,
        margin=dict(t=30, b=10, l=10, r=10),
        xaxis_title="", yaxis_title="Permisos",
        showlegend=False,
    )
    st.plotly_chart(fig1, use_container_width=True)
    st.caption(
        f"Promedio semanal: **{prom:.1f} permisos** — "
        f"La semana con más permisos fue **{sem.loc[sem['cantidad'].idxmax(), 'año_semana']}** "
        f"({max_sem} permisos, marcada en rojo)."
    )

    st.divider()

    # ── 1b. Permisos por mes ──
    st.subheader("Permisos por mes")
    df["mes_label"] = df["fecha"].dt.strftime("%Y-%m")
    mes_g    = df.groupby("mes_label").size().reset_index(name="cantidad")
    prom_mes = mes_g["cantidad"].mean()
    max_mes  = mes_g["cantidad"].max()
    cols_mes = ["#C0392B" if v == max_mes else "#2471D5" for v in mes_g["cantidad"]]

    fig1b = go.Figure(go.Bar(
        x=mes_g["mes_label"], y=mes_g["cantidad"],
        text=mes_g["cantidad"], textposition="outside",
        marker_color=cols_mes,
        hovertemplate="%{x}: %{y} permisos<extra></extra>",
    ))
    fig1b.add_hline(
        y=prom_mes, line_dash="dash", line_color="#7F8C8D",
        annotation_text=f"Promedio: {prom_mes:.1f}",
        annotation_position="top left",
    )
    fig1b.update_layout(
        plot_bgcolor="white", height=270,
        margin=dict(t=30, b=10, l=10, r=10),
        xaxis_title="", yaxis_title="Permisos", showlegend=False,
    )
    st.plotly_chart(fig1b, use_container_width=True)
    mes_pico = mes_g.loc[mes_g["cantidad"].idxmax(), "mes_label"]
    st.caption(
        f"Promedio: **{prom_mes:.1f} permisos/mes** — "
        f"Mes pico: **{mes_pico}** con {max_mes} permisos 🔴"
    )

    st.divider()

    # ── 2. Pareto de motivos ──
    st.subheader("¿Por qué salen? — Motivos principales (Pareto 80%)")
    mc = df["motivo"].value_counts().reset_index()
    mc.columns = ["Motivo", "Cantidad"]
    mc["acum_pct"] = (mc["Cantidad"].cumsum() / mc["Cantidad"].sum() * 100)
    pareto = mc[mc["acum_pct"].shift(1, fill_value=0) < 80].head(8)

    max_mot = pareto["Cantidad"].max()
    colores_mot = ["#C0392B" if v == max_mot else "#2471D5" for v in pareto["Cantidad"]]

    fig2 = go.Figure(go.Bar(
        x=pareto["Cantidad"],
        y=pareto["Motivo"],
        orientation="h",
        text=pareto["Cantidad"],
        textposition="outside",
        marker_color=colores_mot,
        hovertemplate="%{y}: %{x} veces<extra></extra>",
    ))
    fig2.update_layout(
        plot_bgcolor="white", height=max(250, len(pareto) * 40),
        margin=dict(t=10, b=10, l=10, r=60),
        xaxis_title="Cantidad", yaxis_title="",
        yaxis={"categoryorder": "total ascending"},
        showlegend=False,
    )
    st.plotly_chart(fig2, use_container_width=True)
    if not pareto.empty:
        top1 = pareto.iloc[-1]
        st.caption(
            f"**{top1['Motivo']}** es el motivo más frecuente ({top1['Cantidad']} veces). "
            f"Estos {len(pareto)} motivos explican el 80% de los permisos."
        )

    st.divider()

    # ── 3. Duración corregida ──
    st.subheader("¿Cuánto tiempo suelen estar fuera?")

    df_dur = df[df["minutos_reales"].notna() & (df["minutos_reales"] > 0)].copy()

    if not df_dur.empty:
        def cat_dur(m):
            if m < 30:    return "< 30 min"
            elif m < 60:  return "30 – 60 min"
            elif m < 90:  return "1h – 1h 30min"
            elif m < 120: return "1h 30min – 2h"
            else:         return "Más de 2h"

        orden = ["< 30 min", "30 – 60 min", "1h – 1h 30min", "1h 30min – 2h", "Más de 2h"]
        df_dur["rango"] = df_dur["minutos_reales"].apply(cat_dur)
        c = (
            df_dur["rango"]
            .value_counts()
            .reindex(orden, fill_value=0)
            .reset_index()
        )
        c.columns = ["Rango", "Cantidad"]
        c["Pct"] = (c["Cantidad"] / c["Cantidad"].sum() * 100).round(1)

        max_dur = c["Cantidad"].max()
        colores_dur = ["#C0392B" if v == max_dur else "#1B4F9B" for v in c["Cantidad"]]

        cd1, cd2 = st.columns([2, 1])
        with cd1:
            fig3 = go.Figure(go.Bar(
                x=c["Rango"],
                y=c["Cantidad"],
                text=c["Pct"].apply(lambda x: f"{x}%"),
                textposition="outside",
                marker_color=colores_dur,
                hovertemplate="%{x}: %{y} permisos (%{text})<extra></extra>",
            ))
            fig3.update_layout(
                plot_bgcolor="white", height=270,
                margin=dict(t=10, b=10, l=10, r=10),
                xaxis_title="", yaxis_title="Permisos",
                showlegend=False,
            )
            st.plotly_chart(fig3, use_container_width=True)
        with cd2:
            st.dataframe(
                c[["Rango", "Cantidad", "Pct"]].rename(columns={"Pct": "%"}),
                use_container_width=True,
                hide_index=True,
            )
            mayor = c.loc[c["Cantidad"].idxmax()]
            prom_min = df_dur["minutos_reales"].mean()
            st.caption(
                f"Rango más frecuente: **{mayor['Rango']}** ({mayor['Pct']}%).\n\n"
                f"Promedio real de ausencia: **{fmt_dur(prom_min)}**."
            )
    else:
        st.info("No hay datos de duración suficientes para este período.")

    st.divider()

    # ── 4. ¿Compensan? Intención vs Realidad ──
    st.subheader("Compensación: intención vs. realidad")
    st.caption(
        "Izquierda: cuántos eligen compensar al salir. "
        "Derecha: de los que dijeron SI, cuántos realmente lo hicieron."
    )

    ratio = df["compensa"].value_counts().reset_index()
    ratio.columns = ["Compensa", "Cantidad"]
    total_r = ratio["Cantidad"].sum()

    colores_ratio = {"SI": "#1A7A4A", "NO": "#C0392B"}

    fig4 = go.Figure(go.Pie(
        labels=ratio["Compensa"],
        values=ratio["Cantidad"],
        hole=0.5,
        marker_colors=[colores_ratio.get(v, "#95A5A6") for v in ratio["Compensa"]],
        textinfo="label+percent",
        hovertemplate="%{label}: %{value} permisos (%{percent})<extra></extra>",
    ))
    fig4.update_layout(
        height=260, margin=dict(t=30, b=10, l=10, r=10),
        showlegend=False, title_text="Intención al salir",
    )

    # Segundo gráfico: de los que dijeron SI, cuántos tienen al menos 1 compensación registrada
    dijeron_si = df[df["compensa"] == "SI"]["legajo"].unique()
    n_dijeron_si = len(dijeron_si)

    if n_dijeron_si > 0 and not comp_planta.empty:
        compensaron = comp_planta[comp_planta["legajo"].isin(dijeron_si)]["legajo"].unique()
        n_compensaron    = len(compensaron)
        n_no_compensaron = n_dijeron_si - n_compensaron

        labels_r = ["Compensaron", "No compensaron"]
        values_r = [n_compensaron, n_no_compensaron]
        cols_r   = ["#1A7A4A", "#E67E22"]

        fig5 = go.Figure(go.Pie(
            labels=labels_r, values=values_r, hole=0.5,
            marker_colors=cols_r,
            textinfo="label+percent",
            hovertemplate="%{label}: %{value} personas (%{percent})<extra></extra>",
        ))
        fig5.update_layout(
            height=260, margin=dict(t=30, b=10, l=10, r=10),
            showlegend=False, title_text="Realidad (de los que dijeron SI)",
        )

        cr1, cr2 = st.columns(2)
        with cr1:
            st.plotly_chart(fig4, use_container_width=True)
            for _, row in ratio.iterrows():
                pct   = row["Cantidad"] / total_r * 100
                icono = "✅" if row["Compensa"] == "SI" else "❌"
                st.metric(f"{icono} {row['Compensa']}", f"{row['Cantidad']} permisos", f"{pct:.1f}%")
        with cr2:
            st.plotly_chart(fig5, use_container_width=True)
            st.metric("✅ Compensaron efectivamente", f"{n_compensaron} personas",
                      f"{n_compensaron/n_dijeron_si*100:.1f}% de los que dijeron SI")
            st.metric("⚠️ Dijeron SI pero no compensaron aún", f"{n_no_compensaron} personas",
                      f"{n_no_compensaron/n_dijeron_si*100:.1f}% de los que dijeron SI")
    else:
        st.plotly_chart(fig4, use_container_width=True)
        st.info("Aún no hay compensaciones registradas para cruzar con los permisos.")

    st.divider()
    st.caption("Análisis automático basado en los datos de Google Sheets.")

    # ── 5. Horas no trabajadas MOD por día ──────────────────────
    st.divider()
    st.subheader("🔧 Horas no trabajadas — Mano de Obra Directa (MOD)")
    st.caption(
        "Permisos de empleados **HOURLY DIRECT** del sector **COSTURA**. "
        "Métrica clave de calidad de producción — calculada automáticamente."
    )

    # ── Selector de rango de fechas (reemplaza año/mes)
    _mod_col1, _mod_col2 = st.columns(2)
    with _mod_col1:
        _mod_desde = st.date_input(
            "📅 Desde",
            value=date.today().replace(day=1),
            max_value=date.today(),
            format="DD/MM/YYYY",
            key="mod_desde",
        )
    with _mod_col2:
        _mod_hasta = st.date_input(
            "📅 Hasta",
            value=date.today(),
            max_value=date.today(),
            format="DD/MM/YYYY",
            key="mod_hasta",
        )

    if _mod_desde > _mod_hasta:
        st.warning("⚠️ La fecha 'Desde' no puede ser posterior a 'Hasta'.")
        st.stop()

    _mod_label = f"{_mod_desde.strftime('%d/%m/%Y')} — {_mod_hasta.strftime('%d/%m/%Y')}"
    # 0.5h = almuerzo a descontar por día con permisos (pausa de mediodía)
    _ALMUERZO_H = 0.5

    # Cruzar permisos con padrón para obtener clasificacion y sector
    _df_mod = permisos_planta.copy()
    _df_mod["sector_emp"] = _df_mod["legajo"].map(sector_dict).fillna("")
    _df_mod["clasif_emp"] = _df_mod["legajo"].map(clasif_dict).fillna("")

    # Filtrar: solo MOD = HOURLY DIRECT + sector que contenga "COSTURA"
    _df_mod = _df_mod[
        (_df_mod["clasif_emp"].str.upper() == "HOURLY DIRECT") &
        (_df_mod["sector_emp"].str.upper().str.contains("COSTURA", na=False))
    ].copy()

    # Filtrar por rango de fechas
    _df_mod = _df_mod[
        (_df_mod["fecha"].dt.date >= _mod_desde) &
        (_df_mod["fecha"].dt.date <= _mod_hasta)
    ].copy()

    if _df_mod.empty:
        st.info(f"No hay permisos de MOD en {_mod_label}.")
    else:
        _df_mod["fecha_str"] = _df_mod["fecha"].dt.strftime("%d/%m/%Y")
        # Horas brutas por día
        _resumen_dia = (
            _df_mod.groupby("fecha_str")
            .agg(
                hs_brutas=("minutos_reales", lambda x: x.sum() / 60),
                personas=("nombre", "nunique"),
                nombres=("nombre", lambda x: ", ".join(sorted(x.unique())))
            )
            .reset_index()
        )
        # Descontar almuerzo (0.5h) por día — tiempo de pausa de mediodía
        _resumen_dia["Hs. no trabajadas"] = (
            (_resumen_dia["hs_brutas"] - _ALMUERZO_H).clip(lower=0).round(2)
        )
        _resumen_dia = _resumen_dia.rename(columns={
            "fecha_str": "Fecha", "personas": "Personas", "nombres": "Empleados"
        })[["Fecha", "Hs. no trabajadas", "Personas", "Empleados"]].sort_values("Fecha")

        _total_mod = _resumen_dia["Hs. no trabajadas"].sum()

        _tm1, _tm2, _tm3 = st.columns(3)
        _tm1.metric("Total hs. no trabajadas (MOD)", f"{_total_mod:.2f}h")
        _tm2.metric("Días con ausentismo MOD", len(_resumen_dia))
        _tm3.metric("Personas MOD afectadas", _df_mod["nombre"].nunique())

        st.dataframe(_resumen_dia, use_container_width=True, hide_index=True,
                     height=min(420, 45 + len(_resumen_dia) * 35))

        _max_mod = _resumen_dia["Hs. no trabajadas"].max()
        _cols_mod = ["#C0392B" if v == _max_mod else "#E67E22"
                     for v in _resumen_dia["Hs. no trabajadas"]]
        _fig_mod = go.Figure(go.Bar(
            x=_resumen_dia["Fecha"], y=_resumen_dia["Hs. no trabajadas"],
            text=_resumen_dia["Hs. no trabajadas"].apply(lambda x: f"{x:.2f}h"),
            textposition="outside", marker_color=_cols_mod,
            hovertemplate="%{x}<br>%{y:.2f}h no trabajadas<extra></extra>",
        ))
        _fig_mod.update_layout(
            plot_bgcolor="white", height=260,
            margin=dict(t=10, b=10, l=10, r=10),
            xaxis_title="", yaxis_title="Horas no trabajadas", showlegend=False,
        )
        st.plotly_chart(_fig_mod, use_container_width=True)
        st.caption(
            f"Período: {_mod_label}. Solo HOURLY DIRECT de Costura. "
            f"Horas brutas menos {_ALMUERZO_H}h de almuerzo por día. Decimales exactos."
        )

    # ── 5b. MOD que NO compensa — para evaluar prima de producción ──
    st.divider()
    st.subheader("🔧❌ Horas MOD sin compensar")
    st.caption(
        "Permisos de **HOURLY DIRECT** del sector **COSTURA** donde la persona "
        "eligió o le correspondió **NO compensar**. Es tiempo de producción "
        "perdido que no se recupera — dato clave para evaluar prima de producción."
    )

    # Usa el mismo rango _mod_desde/_mod_hasta del bloque MOD de arriba
    _df_mod_nc = permisos_planta.copy()
    _df_mod_nc["sector_emp"] = _df_mod_nc["legajo"].map(sector_dict).fillna("")
    _df_mod_nc["clasif_emp"] = _df_mod_nc["legajo"].map(clasif_dict).fillna("")

    _df_mod_nc = _df_mod_nc[
        (_df_mod_nc["clasif_emp"].str.upper() == "HOURLY DIRECT") &
        (_df_mod_nc["sector_emp"].str.upper().str.contains("COSTURA", na=False)) &
        (_df_mod_nc["compensa"] == "NO") &
        (_df_mod_nc["fecha"].dt.date >= _mod_desde) &
        (_df_mod_nc["fecha"].dt.date <= _mod_hasta)
    ].copy()

    if _df_mod_nc.empty:
        st.success(f"✅ No hay permisos MOD sin compensar en {_mod_label}.")
    else:
        # Horas desde minutos_reales (decimales exactos) o fallback a hora_salida/entrada
        def _calc_hs_mod_nc(row):
            if pd.notna(row["minutos_reales"]) and row["minutos_reales"] > 0:
                return round(row["minutos_reales"] / 60, 2)
            try:
                sal, ent = row["hora_salida"], row["hora_entrada"]
                if isinstance(sal, str) and isinstance(ent, str) and len(sal)==5 and len(ent)==5 and ent!="S/R":
                    h_s, m_s = map(int, sal.split(":"))
                    h_e, m_e = map(int, ent.split(":"))
                    mins = (h_e*60+m_e) - (h_s*60+m_s)
                    return round(mins/60, 2) if mins > 0 else 0.0
            except Exception:
                pass
            return 0.0

        _df_mod_nc["hs_real"] = _df_mod_nc.apply(_calc_hs_mod_nc, axis=1)
        _df_mod_nc["fecha_str"] = _df_mod_nc["fecha"].dt.strftime("%d/%m/%Y")

        # Agrupar por día y descontar almuerzo
        _resumen_mod_nc = (
            _df_mod_nc.groupby("fecha_str")
            .agg(
                total_horas=("hs_real", "sum"),
                personas=("nombre", "nunique"),
                nombres=("nombre", lambda x: ", ".join(sorted(x.unique()))),
            )
            .reset_index()
        )
        _resumen_mod_nc["total_horas"] = (
            (_resumen_mod_nc["total_horas"] - _ALMUERZO_H).clip(lower=0).round(2)
        )
        _resumen_mod_nc = _resumen_mod_nc.rename(columns={
            "fecha_str": "Fecha", "total_horas": "Hs. no compensadas",
            "personas": "Personas", "nombres": "Empleados",
        }).sort_values("Fecha")

        _total_nc = _resumen_mod_nc["Hs. no compensadas"].sum()
        _tmnc1, _tmnc2, _tmnc3 = st.columns(3)
        _tmnc1.metric("Total hs. MOD no compensadas", f"{_total_nc:.2f}h")
        _tmnc2.metric("Días con ausentismo sin compensar", len(_resumen_mod_nc))
        _tmnc3.metric("Personas MOD afectadas", _df_mod_nc["nombre"].nunique())

        st.dataframe(_resumen_mod_nc, use_container_width=True, hide_index=True,
                     height=min(420, 45 + len(_resumen_mod_nc) * 35))

        st.markdown("**Detalle por persona y motivo**")
        _detalle_mod_nc = (
            _df_mod_nc.groupby(["nombre", "motivo"])
            .agg(permisos=("fecha_str", "count"), horas=("hs_real", "sum"))
            .reset_index()
            .sort_values("horas", ascending=False)
        )
        _detalle_mod_nc["horas"] = _detalle_mod_nc["horas"].apply(lambda x: f"{x:.2f}h")
        _detalle_mod_nc.columns = ["Nombre", "Motivo", "Permisos", "Horas"]
        st.dataframe(_detalle_mod_nc, use_container_width=True, hide_index=True)

        st.caption(
            f"Período: {_mod_label}. Solo HOURLY DIRECT de Costura con compensa=NO. "
            f"Descuento de {_ALMUERZO_H}h de almuerzo por día. Decimales exactos."
        )

    # ── 6. (OPCIONAL) Tiempo no productivo acumulado por mes ──
    st.divider()
    st.subheader("⏳ Tiempo no productivo acumulado")
    st.caption(
        "Horas que se deben compensar pero todavía no se recuperaron, agrupadas por mes de origen. "
        "Cada hora pendiente es tiempo de producción no recuperado."
    )

    if not permisos_planta.empty:
        # Horas comprometidas por mes (solo las que eligen compensar)
        # Usamos permisos_planta completo (sin filtro de año) para el histórico total
        p_comp = permisos_planta[permisos_planta["compensa"] == "SI"].copy()
        p_comp["mes_label"] = p_comp["fecha"].dt.strftime("%Y-%m")
        hs_por_mes = p_comp.groupby("mes_label")["horas_redondeadas"].sum().reset_index()
        hs_por_mes.columns = ["Mes", "Horas comprometidas"]

        # Total compensado hasta hoy
        total_comp = compensaciones[compensaciones["planta"] == KEY_PLANTA]["horas_compensadas"].sum()             if not compensaciones.empty else 0
        total_comprometido = hs_por_mes["Horas comprometidas"].sum()
        pendiente_total = max(0, total_comprometido - total_comp)

        # Mostrar métricas clave
        tp1, tp2, tp3 = st.columns(3)
        tp1.metric("Horas comprometidas (total)", f"{total_comprometido:.0f}h")
        tp2.metric("Horas ya recuperadas", f"{total_comp:.0f}h")
        tp3.metric("⚠️ Horas aún sin recuperar", f"{pendiente_total:.0f}h",
                   delta=f"-{pendiente_total:.0f}h vs producción plena", delta_color="inverse")

        # Gráfico de barras por mes
        max_np = hs_por_mes["Horas comprometidas"].max()
        cols_np = ["#C0392B" if v == max_np else "#E67E22" for v in hs_por_mes["Horas comprometidas"]]

        fig_np = go.Figure(go.Bar(
            x=hs_por_mes["Mes"],
            y=hs_por_mes["Horas comprometidas"],
            text=hs_por_mes["Horas comprometidas"].apply(lambda x: f"{x:.0f}h"),
            textposition="outside",
            marker_color=cols_np,
            hovertemplate="%{x}: %{y}h comprometidas<extra></extra>",
        ))
        fig_np.update_layout(
            plot_bgcolor="white", height=270,
            margin=dict(t=30, b=10, l=10, r=10),
            xaxis_title="", yaxis_title="Horas", showlegend=False,
            title_text="Horas comprometidas a compensar por mes de origen",
        )
        st.plotly_chart(fig_np, use_container_width=True)
        st.caption(
            f"De las **{total_comprometido:.0f}h** totales comprometidas, "
            f"se recuperaron **{total_comp:.0f}h** ({total_comp/total_comprometido*100:.1f}% si hubiera datos). "
            f"Quedan **{pendiente_total:.0f}h** sin recuperar."
            if total_comprometido > 0 else "Sin datos suficientes."
        )


# ═══════════════════════════════════════════════════════════════
# CÓMO SE CALCULA — Documentación interna de fórmulas
# ═══════════════════════════════════════════════════════════════
elif pagina == "📖 Cómo se calcula":

    st.title("📖 Cómo se calcula todo en esta app")
    st.caption(
        "Documentación de referencia para entender la lógica de cálculo. "
        "Cualquier persona nueva en el equipo puede leer esto antes de tocar datos."
    )
    st.divider()

    # ── 1. Redondeo de horas ──────────────────────────────────
    with st.expander("⏱ 1. Redondeo de horas — regla de fábrica", expanded=False):
        st.markdown("""
Toda ausencia parcial se convierte a horas enteras con esta regla:

| Minutos reales | Horas redondeadas |
|---|---|
| < 30 min | 0 horas |
| 30 – 89 min | 1 hora |
| 90 – 149 min | 2 horas |
| 150 – 209 min | 3 horas |

**Regla general:** si la fracción sobre la hora entera es ≥ 30 minutos, sube al entero siguiente.
        """)
        st.markdown("**Ejemplos reales:**")
        ej_df = pd.DataFrame({
            "Salida": ["09:50", "09:00", "08:00", "11:00"],
            "Entrada": ["10:40", "10:45", "09:10", "13:24"],
            "Tiempo real": ["50 min", "1h 45min", "1h 10min", "2h 24min"],
            "Horas redondeadas": ["1 hora", "2 horas", "1 hora", "2 horas"],
        })
        st.dataframe(ej_df, use_container_width=True, hide_index=True)

        st.markdown("**Sin retorno (S/R):** se calcula contra el fin de turno (15:00 hs), no contra el final del día.")
        st.code("Minutos ausentes = (15:00 - hora_salida) en minutos\nEjemplo: salida 10:00 → 300 min → 5 horas", language="text")

    # ── 2. Saldo pendiente ────────────────────────────────────
    with st.expander("💰 2. Cálculo de saldo pendiente por persona", expanded=False):
        st.markdown("""
El saldo es **acumulativo** — no se resetea por mes. Una persona puede deber horas
de febrero y compensarlas en mayo. Eso es correcto y esperado.
        """)
        st.code(
            "Debe_total   = Σ horas_redondeadas  (donde compensa = 'SI')\n"
            "Ya_compensó  = Σ horas_compensadas  (todas las fechas registradas)\n"
            "Saldo_actual = Debe_total − Ya_compensó",
            language="text"
        )
        st.caption("Si saldo_actual ≤ 0, la persona NO aparece en el reporte. No hay saldo negativo.")

    # ── 3. Tope anual ──────────────────────────────────────────
    with st.expander("🚦 3. Tope anual de compensación", expanded=False):
        st.markdown(f"""
Cada empleada tiene un límite anual de horas que puede comprometer para compensar.
Al alcanzar ese límite, el sistema **no permite** marcar nuevos permisos como
`compensa=SI` — los fuerza automáticamente a `NO`.
        """)
        st.code(
            f"Tope_anual = {TOPE_HORAS_NORMAL}h   (empleada normal)\n"
            f"Tope_anual = {TOPE_HORAS_LIDER}h  (líderes)\n\n"
            "Consumido_año = Σ horas_redondeadas donde:\n"
            "  • compensa = 'SI'\n"
            "  • año(fecha) = año calendario actual\n\n"
            "Disponible = MAX(0, Tope_anual − Consumido_año)\n"
            "Excedente  = MAX(0, Consumido_año − Tope_anual)",
            language="text"
        )
        st.caption("Año calendario: resetea el 1° de enero de cada año.")

        st.markdown("**Semáforo de estado:**")
        semaforo_df = pd.DataFrame({
            "Estado": ["🟢 OK", "🟡 ATENCIÓN", "🔴 LÍMITE", "🔴 EXCEDE TOPE"],
            "Condición": [
                "Disponible > 2h",
                "Disponible ≤ 2h",
                "Consumido = Tope",
                "Consumido > Tope",
            ],
            "Consecuencia": [
                "Puede seguir compensando con normalidad",
                "Próximo permiso puede cerrar el tope",
                "No puede sumar más compensa=SI",
                "Nuevos permisos: NO automático",
            ],
        })
        st.dataframe(semaforo_df, use_container_width=True, hide_index=True)

        st.markdown("**Líderes con tope de 16h/año (nombres exactos del padrón):**")
        for lider in sorted(LIDERES_SJ):
            st.markdown(f"- {lider}")

    # ── 4. Política de motivos ─────────────────────────────────
    with st.expander("📋 4. Política de motivos — qué puede compensar y qué no", expanded=False):
        st.markdown("**Fábrica San Juan (RR.HH. 020):**")
        sj_si = pd.DataFrame({"Motivo": sorted(MOTIVOS_COMPENSAN_SJ)})
        sj_si["¿Puede compensar?"] = "✅ SÍ — guardia elige SI/NO"
        st.dataframe(sj_si, use_container_width=True, hide_index=True)

        sj_no = [m for m in MOTIVOS_LISTA if m not in MOTIVOS_COMPENSAN_SJ]
        sj_no_df = pd.DataFrame({"Motivo": sj_no})
        sj_no_df["¿Puede compensar?"] = "❌ NO — sistema fuerza NO"
        st.dataframe(sj_no_df, use_container_width=True, hide_index=True)

        st.warning(
            "⚠️ **'Médico turno mañana'** SÍ puede compensar. **'Médico propio'** "
            "(sin especificar turno) NO puede compensar. Es una distinción de política, no del sistema."
        )

        st.markdown("**Casa Central Bs. As. (convenio 036):**")
        bsas_si = pd.DataFrame({"Motivo": sorted(MOTIVOS_COMPENSAN_BSAS)})
        bsas_si["¿Puede compensar?"] = "✅ SÍ — guardia elige SI/NO"
        st.dataframe(bsas_si, use_container_width=True, hide_index=True)

    # ── 5. Cómo leer el reporte de gerencia ────────────────────
    with st.expander("📄 5. Cómo leer el reporte para gerencia", expanded=False):
        reporte_cols_df = pd.DataFrame({
            "Columna": ["Apellido y Nombre", "Clasificación", "Base anual", "Pendiente",
                       "Consumido año", "Disponible", "Estado"],
            "Qué significa": [
                "Empleada. Las líderes tienen 👑 al lado.",
                "Categoría laboral: HOURLY DIRECT, HOURLY INDIRECT, EXEMPT, NON EXEMPT.",
                "Tope máximo anual: 8h normal, 16h líder.",
                "Horas que todavía debe compensar (acumuladas, todas las fechas).",
                "Horas con compensa=SI en el año calendario actual. Resetea 1° de enero.",
                "Cuánto le queda disponible = Base anual − Consumido año (mínimo 0).",
                "Semáforo según el tope (ver sección 3).",
            ],
        })
        st.dataframe(reporte_cols_df, use_container_width=True, hide_index=True)
        st.error(
            "⚠️ **Importante:** el reporte SOLO debe generarse con el botón "
            "'Descargar PNG' o 'Descargar Excel'. No copiar datos a mano a otra planilla — "
            "eso fue lo que causó la omisión de personas en un reporte anterior."
        )

    # ── 6. MOD y tiempo no productivo ──────────────────────────
    with st.expander("🔧 6. Métricas de producción (MOD)", expanded=False):
        st.markdown("""
**Horas no trabajadas MOD:** todos los permisos de empleados `HOURLY DIRECT` del
sector `COSTURA`, sin importar si compensan o no. Mide tiempo físico fuera de la
línea de producción — usa `minutos_reales` sin redondeo.

**Horas MOD sin compensar:** el mismo filtro de MOD (HOURLY DIRECT + Costura) pero
solo donde `compensa = NO`. Es tiempo de producción que **no se recupera nunca** —
insumo directo para evaluar prima de producción.

**Tiempo no productivo acumulado:** horas comprometidas a compensar (`compensa=SI`)
de toda la planta, agrupadas por mes de origen, comparadas contra lo ya recuperado.
No está limitado a MOD — es la foto general de deuda de horas de toda la dotación.
        """)
        st.info(
            "Estas tres métricas son independientes entre sí. Ninguna combina "
            "automáticamente 'MOD' + 'compensa' salvo donde el título lo dice explícitamente."
        )

    st.divider()
    st.caption(
        "Esta documentación vive dentro del código (app.py) y se actualiza junto con la lógica. "
        "Si una fórmula cambia acá, debe reflejarse también en esta página."
    )