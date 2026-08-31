"""
ui/rrhh.py
─────────────────────────────────────────────────────────────────
Panel de RRHH: reporte gerencia, compensaciones, histórico,
corrección de registros y gestión del padrón.
Solo lógica de presentación.
"""
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from datetime import date, datetime
import calendar as _cal

from config import AÑOS, MESES, CLASIF_COLOR, COLORES, MOTIVOS_FUERA_TOPE
from services.permisos_service import (
    calcular_saldos, fmt_horas, generar_id,
    obtener_tope, horas_comprometidas_año,
    validar_compensacion,
)
from repositories.sheets_repo import (
    guardar_compensacion, corregir_permiso,
    marcar_activo, eliminar_empleado_padron,
    verificar_compensacion_duplicada,
)


def _caption(text: str):
    """Caption con mejor contraste y legibilidad."""
    import re
    text_html = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", text)
    st.markdown(
        f'<p style="color:#2C3E50;font-size:0.95rem;margin-top:-4px;">{text_html}</p>',
        unsafe_allow_html=True,
    )


def render(
    gc,
    planta_activa: str,
    key_planta: str,
    padron_planta: pd.DataFrame,
    padron_activos: pd.DataFrame,
    permisos_activos: pd.DataFrame,
    comp_activos: pd.DataFrame,
    compensaciones: pd.DataFrame,
    nombre_a_legajo: dict,
    nombres_lista: list,
    sector_dict: dict,
    clasif_dict: dict,
    planta_dict: dict,
):
    if "San Juan" in planta_activa:
        _titulo = "🟢 RRHH — San Juan"
    elif "Bs." in planta_activa:
        _titulo = "🟢 RRHH — Bs. As."
    else:
        _titulo = "🟢 RRHH — Total Empresa"

    st.title(_titulo)
    st.markdown("Seguimiento de permisos y compensaciones.")
    st.divider()

    # ── Gestión de empleados ──────────────────────────────────
    st.subheader("Gestionar empleados — marcar inactivo o dar de baja")
    with st.expander("👤 Gestionar empleados — marcar inactivo o dar de baja"):
        _caption("Marcá como Inactivo a quien se fue de la empresa. Sus horas "
                  "pendientes desaparecen del reporte pero quedan en el historial.")
        if not padron_planta.empty:
            emp_sel = st.selectbox(
                "Empleado/a",
                ["— Seleccioná —"] + sorted(padron_planta["nombre"].tolist()),
                key="emp_gestion",
            )
            if emp_sel and emp_sel != "— Seleccioná —":
                activo_vals = padron_planta[padron_planta["nombre"] == emp_sel]["activo"].values
                activo_actual = str(activo_vals[0]).upper() if len(activo_vals) > 0 else "SI"
                st.write(f"Estado actual: **{'✅ Activo' if activo_actual == 'SI' else '⛔ Inactivo'}**")
                cg1, cg2 = st.columns(2)
                with cg1:
                    lbl = "⛔ Marcar como Inactivo" if activo_actual == "SI" else "✅ Reactivar"
                    nuevo = "NO" if activo_actual == "SI" else "SI"
                    if st.button(lbl, use_container_width=True, key="btn_activo"):
                        if marcar_activo(gc, emp_sel, nuevo):
                            st.success(f"{'⛔ Inactivo' if nuevo == 'NO' else '✅ Reactivado'}: {emp_sel}")
                            st.rerun()
                with cg2:
                    if st.button("🗑️ Dar de baja del padrón", use_container_width=True,
                                  type="primary", key="btn_baja"):
                        if eliminar_empleado_padron(gc, emp_sel):
                            st.success(f"🗑️ {emp_sel} eliminado.")
                            st.rerun()
        else:
            st.info("No hay empleados para esta planta.")

    st.divider()

    # ── Filtros ───────────────────────────────────────────────
    st.subheader("Seleccionar 'acumulado' para el reporte gerencia:")
    cf1, cf2, cf3 = st.columns([1, 1, 2])
    with cf1:
        año_sel = st.selectbox("Año", AÑOS, index=AÑOS.index(min(date.today().year, max(AÑOS))))
    with cf2:
        mes_sel = st.selectbox(
            "Mes", list(MESES.keys()),
            index=date.today().month - 1,
            format_func=lambda x: MESES[x],
        )
    with cf3:
        modo = st.radio("Vista:", ["Solo este mes", "Saldo acumulado total"], horizontal=True)

    if permisos_activos.empty:
        st.warning("No hay permisos cargados para esta planta.")
        return

    # Filtrar por modo
    if modo == "Solo este mes":
        p_f = permisos_activos[
            (permisos_activos["fecha"].dt.year == año_sel) &
            (permisos_activos["fecha"].dt.month == mes_sel)
        ].copy()
        c_f = comp_activos[
            (comp_activos["fecha_compensacion"].dt.year == año_sel) &
            (comp_activos["fecha_compensacion"].dt.month == mes_sel)
        ].copy() if not comp_activos.empty else comp_activos.copy()
    else:
        p_f = permisos_activos.copy()
        c_f = comp_activos.copy()

    # Saldos
    _ultimo_dia = date(año_sel, mes_sel, _cal.monthrange(año_sel, mes_sel)[1])
    if modo == "Solo este mes":
        saldos_al_cierre = calcular_saldos(
            permisos_activos[
                (permisos_activos["fecha"].dt.year == año_sel) &
                (permisos_activos["fecha"].dt.month == mes_sel)
            ],
            comp_activos,
            hasta_fecha=_ultimo_dia,
        )
    else:
        saldos_al_cierre = calcular_saldos(permisos_activos, comp_activos)

    saldos_actuales = calcular_saldos(permisos_activos, comp_activos)
    comp_total      = c_f["horas_compensadas"].sum() if not c_f.empty and "horas_compensadas" in c_f.columns else 0
    _label_periodo  = f"{MESES[mes_sel]} {año_sel}" if modo == "Solo este mes" else "Total acumulado"

    # ── Métricas ──────────────────────────────────────────────
    st.divider()
    cm1, cm2, cm3, cm4 = st.columns(4)
    cm1.metric(f"📋 Permisos ({_label_periodo})", len(p_f))
    cm2.metric("👥 Con saldo > 0", len(saldos_al_cierre))
    cm3.metric("⏳ Hs. pendientes (hoy)",
               f"{saldos_actuales['saldo'].sum():.0f}h" if not saldos_actuales.empty else "0h")
    cm4.metric(f"✅ Hs. compensadas ({_label_periodo})", f"{comp_total:.0f}h")

    # ── Reporte para Gerencia ─────────────────────────────────
    st.divider()
    st.subheader(f"📄 Reporte para Gerencia — {MESES[mes_sel]} {año_sel}")
    _caption("Quiénes deben compensar horas, base anual, consumido y disponible. "
             "Descargá en PNG o Excel — no copies a mano.")

    if saldos_al_cierre.empty:
        st.success(f"✅ Sin horas pendientes al cierre de {MESES[mes_sel]} {año_sel}.")
    else:
        rep = saldos_al_cierre.copy()
        rep["sector"]        = rep["legajo"].map(sector_dict).fillna("Sin sector")
        rep["clasificacion"] = rep["legajo"].map(clasif_dict).fillna("Sin clasificar")
        rep["es_lider"]      = rep["legajo"].map(
            dict(zip(padron_activos["legajo"], padron_activos["es_lider"]))
        ).fillna("NO")
        # planta por legajo — necesario en la vista "Total Empresa" para
        # que cada persona use el tope correcto (SJ 8/16h vs Bs.As. 10h fijo).
        rep["planta_legajo"] = rep["legajo"].map(planta_dict).fillna(
            key_planta if key_planta != "Total" else "Fábrica"
        )
        rep["base_anual"]    = rep.apply(
            lambda r: obtener_tope(r["es_lider"], r["planta_legajo"]), axis=1
        )
        rep["consumido_año"] = rep["legajo"].apply(
            lambda leg: horas_comprometidas_año(permisos_activos, leg, año_sel)
        )
        rep["disponible"]    = (rep["base_anual"] - rep["consumido_año"]).clip(lower=0)
        rep["excedente"]     = (rep["consumido_año"] - rep["base_anual"]).clip(lower=0)

        def _estado(row):
            if row["excedente"] > 0:   return "🔴 EXCEDE TOPE"
            if row["disponible"] == 0: return "🔴 LÍMITE"
            if row["disponible"] <= 2: return "🟡 ATENCIÓN"
            return "🟢 OK"

        rep["estado"] = rep.apply(_estado, axis=1)
        rep = rep.sort_values(["sector", "clasificacion", "saldo"], ascending=[True, True, False])

        sector_actual, html_rows = None, []
        for _, row in rep.iterrows():
            if row["sector"] != sector_actual:
                sector_actual = row["sector"]
                html_rows.append(
                    f'<tr style="background:#1B4F9B;color:white;font-weight:700;">'
                    f'<td colspan="7" style="padding:6px 10px;">📁 {sector_actual}</td></tr>'
                )
            bg      = CLASIF_COLOR.get(row["clasificacion"], "#F8F9FA")
            nombre_ = row["nombre"] + (" 👑" if row["es_lider"] == "SI" else "")
            html_rows.append(
                f'<tr style="background:{bg};">'
                f'<td style="padding:5px 10px;">{nombre_}</td>'
                f'<td style="padding:5px 10px;color:#555;font-size:0.85rem;">{row["clasificacion"]}</td>'
                f'<td style="text-align:center;padding:5px;">{fmt_horas(row["base_anual"])}</td>'
                f'<td style="text-align:center;padding:5px;font-weight:700;color:#C0392B;">{fmt_horas(row["saldo"])}</td>'
                f'<td style="text-align:center;padding:5px;">{fmt_horas(row["consumido_año"])}</td>'
                f'<td style="text-align:center;padding:5px;color:#1A7A4A;font-weight:700;">{fmt_horas(row["disponible"])}</td>'
                f'<td style="text-align:center;padding:5px;">{row["estado"]}</td>'
                f'</tr>'
            )

        st.markdown(
            f'<style>.rep-t{{width:100%;border-collapse:collapse;font-size:0.85rem;font-family:sans-serif;}}'
            f'.rep-t th{{background:#1B4F9B;color:white;padding:8px;text-align:left;font-size:0.78rem;}}</style>'
            f'<table class="rep-t"><thead><tr>'
            f'<th>Apellido y Nombre</th><th>Clasificación</th>'
            f'<th style="text-align:center">Base</th><th style="text-align:center">Pendiente</th>'
            f'<th style="text-align:center">Consumido {año_sel}</th>'
            f'<th style="text-align:center">Disponible</th><th style="text-align:center">Estado</th>'
            f'</tr></thead><tbody>{"".join(html_rows)}</tbody></table>',
            unsafe_allow_html=True,
        )
        st.caption("👑 = líder (tope 16h/año)")
        st.markdown("<br>", unsafe_allow_html=True)

        n_exc = (rep["excedente"] > 0).sum()
        if n_exc > 0:
            for _, r in rep[rep["excedente"] > 0].iterrows():
                st.error(
                    f"🔴 **{r['nombre']}** — excede el tope de {fmt_horas(r['base_anual'])} "
                    f"en {fmt_horas(r['excedente'])}. A descontar del jornal: **{fmt_horas(r['excedente'])}**."
                )

        # Descargas
        _dl1, _dl2, _dl3 = st.columns(3)
        _png_rep    = rep[["nombre", "sector", "clasificacion", "base_anual",
                           "saldo", "consumido_año", "disponible", "estado"]].copy()
        for _c in ["base_anual", "saldo", "consumido_año", "disponible"]:
            _png_rep[_c] = _png_rep[_c].apply(fmt_horas)
        _png_rep.columns = ["Nombre", "Sector", "Clasificación", "Base",
                            "Pendiente", f"Consumido {año_sel}", "Disponible", "Estado"]

        _row_h = 32
        _fig_png = go.Figure(go.Table(
            header=dict(values=list(_png_rep.columns),
                        fill_color="#1B4F9B",
                        font=dict(color="white", size=11, family="Arial"),
                        align="left", height=36),
            cells=dict(values=[_png_rep[c] for c in _png_rep.columns],
                       fill_color="white",
                       font=dict(color="#222", size=10, family="Arial"),
                       align="left", height=_row_h),
        ))
        _fig_png.update_layout(
            title=dict(
                text=f"GILDAN — Compensación Horaria | {MESES[mes_sel]} {año_sel} | {planta_activa}",
                font=dict(size=13, color="#1B4F9B"), x=0,
            ),
            margin=dict(t=60, b=20, l=10, r=10),
            height=max(220, 60 + len(_png_rep) * _row_h),
            width=1000,
        )
        with _dl1:
            try:
                _img = _fig_png.to_image(format="png", scale=2)
                st.download_button("⬇️ Descargar PNG", data=_img,
                                   file_name=f"reporte_{MESES[mes_sel]}_{año_sel}.png",
                                   mime="image/png", use_container_width=True)
            except Exception:
                st.caption("PNG no disponible (kaleido no instalado).")
        with _dl2:
            try:
                import io as _io
                import openpyxl
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                _CLASIF_COLOR_XL = {
                    "HOURLY DIRECT": "EBF5FB", "HOURLY INDIRECT": "EAF4F4",
                    "EXEMPT": "FEF9E7", "NON EXEMPT": "FDEDEC",
                }
                _ESTADO_COLOR_XL = {
                    "🟢 OK": "D5EFE3", "🟡 ATENCIÓN": "FEF0E0",
                    "🔴 LÍMITE": "FDEDEC", "🔴 EXCEDE TOPE": "FDEDEC",
                }

                _wb_xl = openpyxl.Workbook()
                _ws_xl = _wb_xl.active
                _ws_xl.title = f"Reporte {MESES[mes_sel]} {año_sel}"

                _xl_cols = list(_png_rep.columns)
                _ws_xl.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_xl_cols))
                _ws_xl.cell(row=1, column=1,
                    value=f"GILDAN — Compensación Horaria | {MESES[mes_sel]} {año_sel} | {planta_activa}")
                _ws_xl["A1"].font = Font(bold=True, size=13, color="FFFFFF")
                _ws_xl["A1"].fill = PatternFill("solid", fgColor="1B4F9B")
                _ws_xl.row_dimensions[1].height = 28

                for ci, col_name in enumerate(_xl_cols, 1):
                    cell = _ws_xl.cell(row=2, column=ci, value=col_name)
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                    cell.fill = PatternFill("solid", fgColor="2471D5")
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                _ws_xl.row_dimensions[2].height = 22

                _thin = Side(style="thin", color="CCCCCC")
                _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
                _sector_xl = None
                _xl_row = 3
                for _, xl_r in _png_rep.iterrows():
                    if xl_r["Sector"] != _sector_xl:
                        _sector_xl = xl_r["Sector"]
                        _ws_xl.merge_cells(start_row=_xl_row, start_column=1,
                                            end_row=_xl_row, end_column=len(_xl_cols))
                        _c = _ws_xl.cell(row=_xl_row, column=1, value=f"  {_sector_xl}")
                        _c.font = Font(bold=True, color="FFFFFF", size=10)
                        _c.fill = PatternFill("solid", fgColor="1B4F9B")
                        _xl_row += 1
                    _clasif_fill = _CLASIF_COLOR_XL.get(xl_r.get("Clasificación", ""), "FFFFFF")
                    for ci, col_name in enumerate(_xl_cols, 1):
                        cell = _ws_xl.cell(row=_xl_row, column=ci, value=xl_r[col_name])
                        cell.border = _border
                        cell.alignment = Alignment(horizontal="center" if ci > 2 else "left")
                        cell.font = Font(size=10)
                        if col_name == "Clasificación":
                            cell.fill = PatternFill("solid", fgColor=_clasif_fill)
                        elif col_name == "Estado":
                            cell.fill = PatternFill("solid", fgColor=_ESTADO_COLOR_XL.get(str(xl_r[col_name]), "FFFFFF"))
                            cell.font = Font(bold=True, size=10)
                        elif col_name == "Pendiente":
                            cell.fill = PatternFill("solid", fgColor="FDEDEC")
                            cell.font = Font(color="C0392B", bold=True, size=10)
                        elif col_name == "Disponible":
                            cell.fill = PatternFill("solid", fgColor="E8F5EE")
                            cell.font = Font(color="1A7A4A", bold=True, size=10)
                        else:
                            cell.fill = PatternFill("solid", fgColor="FFFFFF")
                    _xl_row += 1

                for ci, w in enumerate([32, 22, 18, 11, 11, 14, 11, 16], 1):
                    _ws_xl.column_dimensions[get_column_letter(ci)].width = w

                _xl_bytes = _io.BytesIO()
                _wb_xl.save(_xl_bytes)
                st.download_button(
                    "⬇️ Descargar Excel", data=_xl_bytes.getvalue(),
                    file_name=f"reporte_{MESES[mes_sel]}_{año_sel}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as _xl_err:
                st.caption(f"Excel no disponible: {_xl_err}")
        with _dl3:
            st.download_button(
                "⬇️ Descargar CSV",
                _png_rep.to_csv(index=False, encoding="utf-8-sig"),
                file_name=f"reporte_{MESES[mes_sel]}_{año_sel}.csv",
                mime="text/csv", use_container_width=True,
            )

    # ── Reporte NO compensan ──────────────────────────────────
    with st.expander("Ver reporte de permisos sin compensación"):
        _caption("Personas que eligieron NO compensar. Dato para evaluar prima de producción.")
        p_no = p_f[p_f["compensa"] == "NO"].copy() if not p_f.empty else pd.DataFrame()
        if p_no.empty:
            st.success("✅ No hay permisos sin compensación en este período.")
        else:
            p_no["sector_nc"] = p_no["legajo"].map(sector_dict).fillna("Sin sector")
            p_no["clasif_nc"] = p_no["legajo"].map(clasif_dict).fillna("Sin clasificar")
            p_no["fecha_str"] = p_no["fecha"].dt.strftime("%d/%m/%Y")

            def _calc_nc(row):
                if pd.notna(row["minutos_reales"]) and row["minutos_reales"] > 0:
                    return round(row["minutos_reales"] / 60, 2)
                if pd.notna(row["horas_redondeadas"]) and row["horas_redondeadas"] > 0:
                    return float(row["horas_redondeadas"])
                try:
                    sal, ent = row["hora_salida"], row["hora_entrada"]
                    if isinstance(sal, str) and isinstance(ent, str) and ent != "S/R":
                        h_s, m_s = map(int, sal.split(":"))
                        h_e, m_e = map(int, ent.split(":"))
                        return round(((h_e*60+m_e)-(h_s*60+m_s))/60, 2)
                except Exception:
                    pass
                return 0.0

            p_no["hs_real"] = p_no.apply(_calc_nc, axis=1)
            rep_nc = (
                p_no.groupby(["nombre", "sector_nc", "clasif_nc"])
                .agg(permisos=("fecha_str", "count"), horas_no_comp=("hs_real", "sum"))
                .reset_index()
                .sort_values(["sector_nc", "horas_no_comp"], ascending=[True, False])
            )
            rep_nc = rep_nc[rep_nc["horas_no_comp"] > 0].copy()
            rep_nc["horas_no_comp"] = rep_nc["horas_no_comp"].apply(fmt_horas)
            rep_nc.columns = ["Nombre", "Sector", "Clasificación", "Permisos", "Hs. no compensadas"]
            st.dataframe(rep_nc, use_container_width=True, hide_index=True)

    # ── Histórico de compensaciones ───────────────────────────
    st.divider()
    st.subheader("📜 Histórico de compensaciones")
    _caption("Consultá todas las compensaciones de una persona y su situación de tope.")
    hist_sel = st.selectbox(
        "Empleado/a",
        ["— Seleccioná —"] + nombres_lista,
        key="hist_comp_sel",
    )
    if hist_sel and hist_sel != "— Seleccioná —":
        _leg_h = nombre_a_legajo.get(hist_sel, "")
        _hc    = comp_activos[comp_activos["legajo"] == _leg_h].copy() if not comp_activos.empty else pd.DataFrame()
        _hp    = permisos_activos[
            (permisos_activos["legajo"] == _leg_h) &
            (permisos_activos["compensa"] == "SI")
        ].copy() if not permisos_activos.empty else pd.DataFrame()
        _lider  = padron_activos[padron_activos["legajo"] == _leg_h]["es_lider"].values
        _lider  = _lider[0] if len(_lider) > 0 else "NO"
        _planta_h = planta_dict.get(_leg_h, key_planta if key_planta != "Total" else "Fábrica")
        _tope   = obtener_tope(_lider, _planta_h)
        _cons   = horas_comprometidas_año(permisos_activos, _leg_h, date.today().year)

        hh1, hh2, hh3 = st.columns(3)
        hh1.metric("Tipo", "👑 Líder" if _lider == "SI" else "Empleada")
        hh1.caption(f"Tope: {_tope:.0f}h/año")
        hh2.metric(f"Consumido {date.today().year}", fmt_horas(_cons))
        hh3.metric("Compensado histórico total",
                   fmt_horas(_hc["horas_compensadas"].sum()) if not _hc.empty else "0h")

        st.markdown("**Permisos que generaron deuda**")
        if _hp.empty:
            st.caption("Sin permisos con compensa=SI.")
        else:
            _hp_s = _hp[["fecha", "motivo", "horas_redondeadas"]].copy()
            _hp_s["fecha"] = _hp_s["fecha"].dt.strftime("%d/%m/%Y")
            _hp_s["horas_redondeadas"] = _hp_s["horas_redondeadas"].apply(fmt_horas)
            _hp_s.columns = ["Fecha", "Motivo", "Horas"]
            st.dataframe(_hp_s.sort_values("Fecha"), use_container_width=True, hide_index=True)

        st.markdown("**Compensaciones registradas**")
        if _hc.empty:
            st.caption("Sin compensaciones registradas.")
        else:
            _hc_s = _hc[["fecha_compensacion", "horas_compensadas", "observacion", "registrado_por"]].copy()
            _hc_s["fecha_compensacion"] = _hc_s["fecha_compensacion"].dt.strftime("%d/%m/%Y")
            _hc_s["horas_compensadas"]  = _hc_s["horas_compensadas"].apply(fmt_horas)
            _hc_s.columns = ["Fecha", "Horas", "Observación", "Registrado por"]
            st.dataframe(_hc_s.sort_values("Fecha"), use_container_width=True, hide_index=True)

    # ── Corrección de registros ───────────────────────────────
    st.divider()
    st.subheader("🛠️ Corregir o anular un permiso")
    _caption("Corregí un permiso cargado por error. Queda en el historial con la razón del cambio.")
    with st.expander("Abrir corrector de registros"):
        edit_nombre = st.selectbox("Empleado/a", ["— Seleccioná —"] + nombres_lista, key="edit_n")
        if edit_nombre and edit_nombre != "— Seleccioná —":
            _leg_e = nombre_a_legajo.get(edit_nombre, "")
            _pe    = permisos_activos[permisos_activos["legajo"] == _leg_e].copy() if not permisos_activos.empty else pd.DataFrame()
            if not _pe.empty:
                _pe["_label"] = (
                    _pe["fecha"].dt.strftime("%d/%m/%Y") + " | " +
                    _pe["motivo"] + " | " + _pe["compensa"] + " | " +
                    _pe["horas_redondeadas"].apply(fmt_horas)
                )
                reg_sel = st.selectbox("Registro", ["— Seleccioná —"] + _pe["_label"].tolist(), key="edit_r")
                if reg_sel and reg_sel != "— Seleccioná —":
                    _id_sel = _pe[_pe["_label"] == reg_sel]["id"].values[0]
                    st.info(f"ID: **{_id_sel}**")
                    accion  = st.radio("Acción", ["Cambiar compensa=SI a NO", "Cambiar compensa=NO a SI", "Anular registro"], horizontal=True, key="edit_acc")
                    razon   = st.text_input("Razón del cambio *", key="edit_raz")
                    quien   = st.text_input("Tu nombre *", key="edit_qui")
                    if st.button("Aplicar corrección", type="primary", key="btn_editar"):
                        if not razon.strip():
                            st.error("❌ La razón es obligatoria.")
                        elif not quien.strip():
                            st.error("❌ Falta tu nombre.")
                        else:
                            if corregir_permiso(gc, _id_sel, accion, razon.strip(), quien.strip()):
                                st.success("✅ Corrección aplicada.")
                            else:
                                st.error("❌ No se encontró el registro. Recargá la página.")
            else:
                st.info("Sin permisos registrados.")

    # ── Registrar compensación ────────────────────────────────
    st.divider()
    st.subheader("✏️ Registrar compensación")
    _caption("Cuando alguien se queda horas extra para compensar, registralo acá.")

    with st.form("form_comp"):
        nom_c_sel = st.selectbox("Empleado/a", ["— Seleccioná —"] + nombres_lista)
        nom_c     = nom_c_sel if nom_c_sel != "— Seleccioná —" else ""
        leg_c     = nombre_a_legajo.get(nom_c, "") if nom_c else ""

        if nom_c and not saldos_actuales.empty:
            saldo_a = saldos_actuales[saldos_actuales["nombre"] == nom_c]["saldo"].sum()
            if saldo_a > 0:
                st.info(f"Saldo pendiente de **{nom_c}**: **{fmt_horas(saldo_a)}**")
            else:
                st.success(f"**{nom_c}** no tiene horas pendientes.")

        cc3, cc4 = st.columns(2)
        with cc3:
            fecha_comp = st.date_input("Fecha en que compensó", value=date.today(),
                                        max_value=date.today(), format="DD/MM/YYYY")
        with cc4:
            hs_comp = st.number_input("Horas compensadas", min_value=0.5, max_value=8.0, value=1.0, step=0.5)

        obs      = st.text_input("Observación (opcional)")
        registra = st.text_input("Tu nombre *")

        if st.form_submit_button("✅ REGISTRAR COMPENSACIÓN", use_container_width=True, type="primary"):
            errores = validar_compensacion(nom_c, registra)
            if verificar_compensacion_duplicada(comp_activos, leg_c, fecha_comp):
                errores.append(
                    f"{nom_c} ya tiene una compensación registrada el "
                    f"{fecha_comp.strftime('%d/%m/%Y')}. No se puede compensar dos veces el mismo día."
                )
            if errores:
                for e in errores:
                    st.error(f"❌ {e}")
            else:
                try:
                    guardar_compensacion(gc, {
                        "id":                generar_id("C"),
                        "fecha_compensacion": fecha_comp.strftime("%Y-%m-%d"),
                        "legajo":            leg_c,
                        "nombre":            nom_c,
                        "horas_compensadas": hs_comp,
                        "observacion":       obs,
                        "registrado_por":    registra.strip(),
                        "planta":            key_planta,
                        "timestamp":         datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    })
                    st.success(f"✅ {nom_c} — {fmt_horas(hs_comp)} el {fecha_comp.strftime('%d/%m/%Y')}")
                except Exception as e:
                    st.error(f"❌ Error: {e}")

    # ── Detalle del período ───────────────────────────────────
    st.divider()
    st.subheader(f"🔍 Detalle de permisos del período")
    if not p_f.empty:
        det = p_f[[
            "fecha", "legajo", "nombre", "hora_salida", "hora_entrada",
            "sin_retorno", "motivo", "compensa", "horas_redondeadas", "minutos_reales"
        ]].copy()
        det["fecha"] = det["fecha"].dt.strftime("%d/%m/%Y")
        det["minutos_reales"] = det["minutos_reales"].apply(
            lambda m: f"{int(m)} min" if pd.notna(m) and m > 0 else "—"
        )
        det.columns = ["Fecha", "Legajo", "Nombre", "Salida", "Entrada",
                       "S/R", "Motivo", "Compensa", "Hs. (redondeadas)", "Minutos reales"]
        _caption("Revisá la columna Minutos reales si sospechás que el redondeo no fue correcto.")
        st.dataframe(det, use_container_width=True, hide_index=True)
    else:
        st.info(f"No hay permisos en este período.")